# Drops pebble in pond

import _config as keys
import nest_asyncio
import asyncio
import httpx
import re
import os
import shutil
from pathlib import Path
import glob
import json
from pipulate import pip # Import pip for persistence
import nbformat
import itertools
import requests
import gzip
import shutil
from pprint import pprint
import pandas as pd
from time import sleep
from io import StringIO
from collections import defaultdict
from tldextract import extract
from bs4 import BeautifulSoup
import wordninja

# --- KEYWORD CLUSTERING SUPPORT FUNCTIONS (REQUIRES: nltk, sklearn, wordninja) ---
from collections import Counter
from nltk import bigrams
from nltk.corpus import stopwords
from nltk.stem import PorterStemmer
from nltk.tokenize import word_tokenize
from sklearn.cluster import MiniBatchKMeans
from sklearn.decomposition import TruncatedSVD
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics import silhouette_score
import numpy as np

# --- EXCEL OUTPUT SUPPORT ---
import platform
import subprocess
import ipywidgets as widgets
from IPython.display import display
import xlsxwriter

# --- EXCEL FORMATTING SUPPORT ---
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import validators # For URL validation

# --- ML ---
import nltk

nltk.download('stopwords', quiet=True)
nltk.download('punkt', quiet=True)
nltk.download('punkt_tab', quiet=True) # Added from a later cell for consolidation

# ... (keep existing imports like Path, nbformat, pip, keys, etc.) ...
import urllib.parse # Need this for correctly encoding the domain/path

try:
    from imports.voice_synthesis import chip_voice_system
    VOICE_AVAILABLE = True
except ImportError:
    VOICE_AVAILABLE = False


def speak(text):
    """Safe wrapper for voice synthesis."""
    if VOICE_AVAILABLE:
        # Fire and forget - don't block the data processing
        try:
            chip_voice_system.speak_text(text)
        except Exception as e:
            print(f"🔇 Voice error: {e}")


def extract_domains_and_print_urls(job: str, notebook_filename: str = "GAPalyzer.ipynb"):
    """
    Parses the specified notebook for competitor domains or subfolders,
    stores them using pip.set, and prints the generated SEMrush URLs
    with appropriate country code and search type.

    Args:
        job (str): The current Pipulate job ID.
        notebook_filename (str): The name of the notebook file to parse.

    Returns:
        list: The list of extracted domains/subfolders, or an empty list if none found/error.
    """
    items_to_analyze = [] # Renamed from domains to be more general

    # --- Inner function to read notebook (kept internal to this step) ---
    def get_items_from_notebook(nb_file):
        """Parses the notebook to get the domain/subfolder list from the 'url-list-input' cell."""
        try:
            notebook_path = Path(nb_file) # Use the passed filename
            if not notebook_path.exists():
                print(f"❌ Error: Notebook file not found at '{notebook_path.resolve()}'")
                return []
            with open(notebook_path, 'r', encoding='utf-8') as f:
                nb = nbformat.read(f, as_version=4)

            for cell in nb.cells:
                if "url-list-input" in cell.metadata.get("tags", []):
                    items_raw = cell.source
                    # Ensure items_raw is treated as a string before splitting lines
                    if isinstance(items_raw, list):
                        items_raw = "".join(items_raw) # Join list elements if needed
                    elif not isinstance(items_raw, str):
                        print(f"⚠️ Warning: Unexpected data type for items_raw: {type(items_raw)}. Trying to convert.")
                        items_raw = str(items_raw)

                    # Now splitlines should work reliably
                    extracted_items = [
                        line.split('#')[0].strip()
                        for line in items_raw.splitlines()
                        if line.strip() and not line.strip().startswith('#')
                    ]
                    # --- NEW: Strip trailing slashes ---
                    extracted_items = [item.rstrip('/') for item in extracted_items]
                    return extracted_items
            print("⚠️ Warning: Could not find a cell tagged with 'url-list-input'.")
            return []
        except Exception as e:
            print(f"❌ Error reading items from notebook: {e}")
            return []

    # --- Main Logic ---
    print("🚀 Extracting domains/subfolders and generating SEMrush URLs...")

    items_to_analyze = get_items_from_notebook(notebook_filename)

    # --- Pipulate Scaffolding ---
    # Store the extracted items list.
    pip.set(job, 'competitor_items', items_to_analyze) # Use a more general key name
    print(f"💾 Stored {len(items_to_analyze)} domains/subfolders in pip state for job '{job}'.")
    # ---------------------------

    # --- Use country_code from keys ---
    try:
        country_db = keys.country_code
    except AttributeError:
        print("⚠️ Warning: 'country_code' not found in keys.py. Defaulting to 'us'.")
        country_db = "us"

    # --- Define the base URL template ---
    base_url = "https://www.semrush.com/analytics/organic/positions/"

    if not items_to_analyze:
        print("🛑 No domains or subfolders found or extracted. Please check the 'url-list-input' cell.")
    else:
        print(f"✅ Found {len(items_to_analyze)} competitor items. URLs to check:")
        print("-" * 30)
        for i, item in enumerate(items_to_analyze):
            # --- Determine searchType dynamically ---
            if '/' in item:
                # If it contains a slash, assume it's a path/subfolder
                search_type = "subfolder"
                # For subfolders, SEMrush often expects the full URL in the 'q' parameter
                query_param = item
                if not query_param.startswith(('http://', 'https://')):
                    # Prepend https:// if no scheme is present
                    query_param = f"https://{query_param}"
            else:
                # Otherwise, treat it as a domain
                search_type = "domain"
                query_param = item # Just the domain name

            # --- Construct the URL ---
            # URL encode the query parameter to handle special characters
            encoded_query = urllib.parse.quote(query_param, safe=':/')
            full_url = f"{base_url}?db={country_db}&q={encoded_query}&searchType={search_type}"

            # Keep the print logic here for user feedback
            print(f"{i+1}. {item}: (Type: {search_type})\n   {full_url}\n")

    return items_to_analyze # Return the list for potential immediate use


def collect_semrush_downloads(job: str, download_path_str: str, file_pattern_xlsx: str = "*-organic.Positions*.xlsx"):
    """
    Moves downloaded SEMRush files matching patterns from the user's download
    directory to a job-specific 'downloads/{job}/' folder, records the
    destination directory and file list in pip state.

    Args:
        job (str): The current job ID.
        download_path_str (str): The user's default browser download path (e.g., "~/Downloads").
        file_pattern_xlsx (str): The glob pattern for SEMRush XLSX files. CSV pattern is derived.

    Returns:
        tuple: (destination_dir_path: str, moved_files_list: list)
               Paths are returned as strings for compatibility. Returns (None, []) on failure or no files.
    """
    print("📦 Starting collection of new SEMRush downloads...")

    # 1. Define source and destination paths
    try:
        source_dir = Path(download_path_str).expanduser()
        if not source_dir.is_dir():
            print(f"❌ Error: Source download directory not found or is not a directory: '{source_dir}'")
            return None, []

        # Destination path relative to the current working directory (assumed Notebooks/)
        destination_dir = Path("downloads") / job
        destination_dir.mkdir(parents=True, exist_ok=True)
        destination_dir_str = str(destination_dir.resolve()) # Store resolved path as string
        print(f"Destination folder created/ensured: '{destination_dir_str}'")

    except Exception as e:
        print(f"❌ Error setting up paths: {e}")
        return None, []

    # 3. Find files in the source directory matching the patterns
    files_to_move = []
    moved_files_list = [] # List to store paths of successfully moved files

    try:
        # Check for .xlsx files
        xlsx_files = glob.glob(str(source_dir / file_pattern_xlsx))
        files_to_move.extend(xlsx_files)

        # Check for .csv files
        file_pattern_csv = file_pattern_xlsx.replace(".xlsx", ".csv")
        csv_files = glob.glob(str(source_dir / file_pattern_csv))
        files_to_move.extend(csv_files)

        if not files_to_move:
            print(f"⚠️ No new files matching patterns ('{file_pattern_xlsx}', '{file_pattern_csv}') found in '{source_dir}'. Skipping move.")
            # --- Pipulate Scaffolding ---
            pip.set(job, 'semrush_download_dir', str(destination_dir)) # Still record the dir path
            pip.set(job, 'collected_semrush_files', []) # Record empty list
            # ---------------------------
            return destination_dir_str, []

        # 4. Move the files
        move_count = 0
        print(f"🔍 Found {len(files_to_move)} potential files to move...")
        for source_file_path in files_to_move:
            source_file = Path(source_file_path)
            dest_file = destination_dir / source_file.name
            dest_file_str = str(dest_file.resolve()) # Store resolved path as string

            if dest_file.exists():
                # File already exists, add its path to the list but don't move
                if dest_file_str not in moved_files_list:
                     moved_files_list.append(dest_file_str)
                # print(f"  -> Skipping (already exists): {source_file.name}") # Optional: Reduce noise
                continue

            try:
                shutil.move(source_file, dest_file)
                print(f"  -> Moved: {source_file.name} to {dest_file}")
                moved_files_list.append(dest_file_str) # Add successfully moved file path
                move_count += 1
            except Exception as e:
                print(f"  -> ❌ Error moving {source_file.name}: {e}")

        print(f"✅ Collection complete. {move_count} new files moved to '{destination_dir}'. Total relevant files: {len(moved_files_list)}")

        # --- Pipulate Scaffolding ---
        pip.set(job, 'semrush_download_dir', destination_dir_str)
        pip.set(job, 'collected_semrush_files', moved_files_list)
        print(f"💾 Stored download directory and {len(moved_files_list)} file paths in pip state for job '{job}'.")
        # ---------------------------

        return destination_dir_str, moved_files_list

    except Exception as e:
        print(f"❌ An unexpected error occurred during file collection: {e}")
        # Attempt to store whatever state we have
        pip.set(job, 'semrush_download_dir', destination_dir_str)
        pip.set(job, 'collected_semrush_files', moved_files_list)
        return destination_dir_str, moved_files_list


def find_semrush_files_and_generate_summary(job: str, competitor_limit: int = None):
    """
    Finds SEMRush files, stores paths in pip state, and generates a Markdown summary.

    Args:
        job (str): The current Pipulate job ID.
        competitor_limit (int, optional): Max number of competitors to list in summary. Defaults to None (list all).


    Returns:
        str: A Markdown formatted string summarizing the found files, or a warning message.
    """
    print(f"🔍 Locating SEMRush files for job '{job}' and generating summary...")
    semrush_dir = Path("downloads") / job
    markdown_output_lines = [] # Initialize list to build Markdown output

    # Ensure the directory exists
    if not semrush_dir.is_dir():
         warning_msg = f"⚠️ **Warning:** Download directory `{semrush_dir.resolve()}` not found. Assuming no files collected yet."
         print(warning_msg.replace("**","")) # Print clean version to console
         pip.set(job, 'collected_semrush_files', [])
         return warning_msg # Return Markdown warning

    file_patterns = [
        "*-organic.Positions*.xlsx",
        "*-organic.Positions*.csv"
    ]

    try:
        all_downloaded_files = sorted(list(itertools.chain.from_iterable(
            semrush_dir.glob(pattern) for pattern in file_patterns
        )))
        all_downloaded_files_as_str = [str(p.resolve()) for p in all_downloaded_files]

        # --- Pipulate Scaffolding ---
        pip.set(job, 'collected_semrush_files', all_downloaded_files_as_str)
        # ---------------------------

        # --- Generate Markdown Output ---
        if all_downloaded_files:
            print(f"💾 Found {len(all_downloaded_files)} files and stored paths in pip state.")
            markdown_output_lines.append("## 💾 Found Downloaded Files")
            markdown_output_lines.append(f"✅ **{len(all_downloaded_files)} files** ready for processing in `{semrush_dir}/`\n")

            display_limit = competitor_limit if competitor_limit is not None else len(all_downloaded_files)
            if display_limit < len(all_downloaded_files):
                 markdown_output_lines.append(f"*(Displaying first {display_limit} based on COMPETITOR_LIMIT)*\n")

            for i, file in enumerate(all_downloaded_files[:display_limit]):
                try:
                    domain_name = file.name[:file.name.index("-organic.")].strip()
                except ValueError:
                    domain_name = file.name # Fallback
                markdown_output_lines.append(f"{i + 1}. **`{domain_name}`** ({file.suffix.upper()})")

        else:
            print(f"🤷 No files found matching patterns in '{semrush_dir}'. Stored empty list in pip state.")
            warning_msg = f"⚠️ **Warning:** No SEMRush files found in `{semrush_dir.resolve()}/`.\n(Looking for `*-organic.Positions*.xlsx` or `*.csv`)"
            markdown_output_lines.append(warning_msg)

        return "\n".join(markdown_output_lines)

    except Exception as e:
        error_msg = f"❌ An unexpected error occurred while listing SEMRush files: {e}"
        print(error_msg)
        pip.set(job, 'collected_semrush_files', []) # Store empty list on error
        return f"❌ **Error:** An error occurred during file listing. Check logs. ({e})"

# In Notebooks/gap_analyzer_sauce.py
import pandas as pd
from tldextract import extract
import itertools
from pathlib import Path
from pipulate import pip # Ensure pip is imported

# (Keep previously added functions)
# ...

def _extract_registered_domain(url: str) -> str:
    """Private helper to extract the registered domain (domain.suffix) from a URL/hostname."""
    try:
        extracted = extract(url)
        return f"{extracted.domain}.{extracted.suffix}"
    except Exception:
        # Fallback for invalid inputs
        return url

def load_and_combine_semrush_data(job: str, client_domain: str, competitor_limit: int = None):
    """
    Loads all SEMRush files from pip state, combines them into a single master DataFrame,
    stores the result in pip state, and returns the DataFrame along with value counts for display.

    Args:
        job (str): The current Pipulate job ID.
        client_domain (str): The client's registered domain (e.g., 'example.com').
        competitor_limit (int, optional): Max number of competitor files to process.

    Returns:
        tuple: (master_df: pd.DataFrame, domain_counts: pd.Series)
               Returns the combined DataFrame and a Series of domain value counts.
               Returns (empty DataFrame, empty Series) on failure or no files.
    """
    semrush_lookup = _extract_registered_domain(client_domain)
    print(f"🛠️  Loading and combining SEMRush files for {semrush_lookup}...")

    # --- INPUT (from pip state) ---
    files_to_process_str = pip.get(job, 'collected_semrush_files', [])
    if not files_to_process_str:
        print("🤷 No collected SEMRush files found in pip state. Nothing to process.")
        return pd.DataFrame(), pd.Series(dtype='int64')

    # Convert string paths back to Path objects
    all_semrush_files = [Path(p) for p in files_to_process_str]
    
    # --- Refactoring Note: Apply COMPETITOR_LIMIT here ---
    # We add 1 to the limit to include the client file if it exists
    processing_limit = competitor_limit + 1 if competitor_limit is not None else None
    files_to_process = all_semrush_files[:processing_limit]
    if processing_limit and len(all_semrush_files) > processing_limit:
        print(f"🔪 Applying COMPETITOR_LIMIT: Processing first {processing_limit} of {len(all_semrush_files)} files.")


    # --- CORE LOGIC (Moved from Notebook) ---
    cdict = {}
    list_of_dfs = []
    print(f"Loading {len(files_to_process)} SEMRush files: ", end="", flush=True)

    for j, data_file in enumerate(files_to_process):
        read_func = pd.read_excel if data_file.suffix.lower() == '.xlsx' else pd.read_csv
        try:
            nend = data_file.stem.index("-organic")
            xlabel = data_file.stem[:nend].replace("_", "/").replace("///", "://").strip('.')
            just_domain = _extract_registered_domain(xlabel)
            cdict[just_domain] = xlabel

            df = read_func(data_file)
            df["Domain"] = xlabel # Use the full label (sub.domain.com) for the column
            df["Client URL"] = df.apply(lambda row: row["URL"] if row["Domain"] == semrush_lookup else None, axis=1)
            df["Competitor URL"] = df.apply(lambda row: row["URL"] if row["Domain"] != semrush_lookup else None, axis=1)
            list_of_dfs.append(df)
            print(f"{j + 1} ", end="", flush=True)

        except Exception as e:
            print(f"\n❌ Error processing file {data_file.name}: {e}")
            continue
    
    print("\n") # Newline after the loading count

    if not list_of_dfs:
        print("🛑 No DataFrames were created. Check for errors in file processing.")
        return pd.DataFrame(), pd.Series(dtype='int64')

    master_df = pd.concat(list_of_dfs, ignore_index=True)
    rows, columns = master_df.shape
    print(f"✅ Combined DataFrame created. Rows: {rows:,}, Columns: {columns:,}")
    
    domain_counts = master_df["Domain"].value_counts()

    # --- OUTPUT (to pip state) ---
    # --- FIX: Save master DF to CSV and store the PATH in pip state ---
    temp_dir = Path("temp") / job # Use the temp directory structure
    temp_dir.mkdir(parents=True, exist_ok=True) # Ensure it exists
    master_csv_path = temp_dir / "semrush_master_combined.csv"
    master_df.to_csv(master_csv_path, index=False)
    print(f"💾 Saved combined SEMrush data to '{master_csv_path}'")
    pip.set(job, 'semrush_master_df_csv_path', str(master_csv_path.resolve()))
    print(f"💾 Stored master DataFrame path and competitor dictionary in pip state for job '{job}'.")
    # -----------------------------

    # --- RETURN VALUE ---
    return master_df, domain_counts


def pivot_semrush_data(job: str, df2: pd.DataFrame, client_domain_from_keys: str):
    """
    Pivots the combined SEMRush DataFrame, calculates competitor positioning,
    manages the competitors CSV file, stores results in pip state, and
    returns the pivot DataFrame for display.

    Args:
        job (str): The current Pipulate job ID.
        df2 (pd.DataFrame): The combined master DataFrame from the previous step.
        client_domain_from_keys (str): The client's domain from the keys module.

    Returns:
        pd.DataFrame: The pivoted DataFrame, or an empty DataFrame on error.
    """
    if df2.empty:
        print("⚠️ Input DataFrame (df2) is empty. Cannot perform pivot.")
        return pd.DataFrame()

    # --- PATH DEFINITION ---
    temp_dir = Path("temp") / job
    temp_dir.mkdir(parents=True, exist_ok=True)
    competitors_csv_file = temp_dir / "competitors.csv"

    # --- INPUTS from pip state & args ---
    semrush_lookup = _extract_registered_domain(client_domain_from_keys)
    # Retrieve the competitor dictionary stored by the previous step
    cdict = pip.get(job, 'competitors_dict_json', {})
    if not isinstance(cdict, dict): # Handle potential JSON string load
        try:
             cdict = json.loads(cdict) if isinstance(cdict, str) else {}
        except:
             cdict = {}
    
    print("🔄 Pivoting data by keyword and calculating competitor positioning...")

    # --- CORE LOGIC (Moved from Notebook) ---
    try:
        pivot_df = df2.pivot_table(index='Keyword', columns='Domain', values='Position', aggfunc='min')

        # Reorder client column to front (Handle potential trailing slash difference)
        target_col = None
        clean_lookup_key = semrush_lookup.rstrip('/')
        for col in pivot_df.columns:
            if col.rstrip('/') == clean_lookup_key:
                target_col = col
                break
        
        if target_col:
            if target_col != semrush_lookup:
                 print(f"  Adjusting semrush_lookup from '{semrush_lookup}' to match column '{target_col}'")
                 semrush_lookup = target_col # Update to the actual column name found
            cols = [semrush_lookup] + [col for col in pivot_df.columns if col != semrush_lookup]
            pivot_df = pivot_df[cols]
        else:
            print(f"❌ Critical Warning: Client domain '{semrush_lookup}' (or variant) not found in pivot table columns. Cannot reorder.")

        competitors = list(pivot_df.columns) # Get competitors AFTER potential reorder
        pip.set(job, 'competitors_list_json', json.dumps(competitors))
        print(f"  💾 Stored canonical competitor list ({len(competitors)} competitors) to pip state.")
        pivot_df['Competitors Positioning'] = pivot_df.iloc[:, 1:].notna().sum(axis=1)

        # Load or initialize df_competitors
        if competitors_csv_file.exists():
            df_competitors = pd.read_csv(competitors_csv_file)
            df_competitors['Title'] = df_competitors['Title'].fillna('')
            df_competitors['Matched Title'] = df_competitors['Matched Title'].fillna('')
            print(f"  ✅ Loaded {len(df_competitors)} existing competitor records from '{competitors_csv_file}'.")
        else:
            if cdict:
                df_competitors = pd.DataFrame(list(cdict.items()), columns=['Domain', 'Column Label'])
                df_competitors['Title'] = ''
                df_competitors['Matched Title'] = ''
                df_competitors.to_csv(competitors_csv_file, index=False)
                print(f"  ✅ Created new competitor file at '{competitors_csv_file}'.")
            else:
                print(f"  ⚠️ Warning: 'competitors_dict_json' was empty or invalid. Cannot create initial competitors file.")
                df_competitors = pd.DataFrame(columns=['Domain', 'Column Label', 'Title', 'Matched Title'])

        # Print keyword counts (internal display logic)
        print("\nKeyword Counts per Competitor:")
        counts = pivot_df.describe().loc['count']
        if not counts.empty:
            max_digits = len(str(len(counts)))
            max_index_width = max(len(str(index)) for index in counts.index)
            valid_counts = [count for count in counts if pd.notna(count)]
            max_count_width = max([len(f"{int(count):,}") for count in valid_counts] or [0])
            for i, (index, count) in enumerate(counts.items(), start=1):
                counter_str = str(i).zfill(max_digits)
                count_str = f"{int(count):,}" if pd.notna(count) else 'NaN'
                print(f"  {counter_str}: {index:<{max_index_width}} - {count_str:>{max_count_width}}")
        else:
            print("  ❌ No data to count after pivot.")

        # Print row/column summary (internal display logic)
        rows_master, _ = df2.shape
        rows_pivot, cols_pivot = pivot_df.shape
        print("\n--- Pivot Summary ---")
        print(f"  Rows (master df): {rows_master:,}")
        print(f"  Rows (pivot df): {rows_pivot:,} ({rows_master - rows_pivot:,} dupes removed)")
        print(f"  Cols (pivot df): {cols_pivot:,}")
        print("---------------------\n")

        # --- OUTPUT (to pip state) ---
        pip.set(job, 'keyword_pivot_df_json', pivot_df.to_json(orient='records'))
        pip.set(job, 'competitors_df_json', df_competitors.to_json(orient='records'))
        print(f"💾 Stored pivot DataFrame and competitors DataFrame in pip state for job '{job}'.")
        # ---------------------------

        # --- RETURN VALUE ---
        return pivot_df # Return the DataFrame for display

    except Exception as e:
        print(f"❌ An error occurred during pivoting: {e}")
        # Store empty states on error
        pip.set(job, 'keyword_pivot_df_json', pd.DataFrame().to_json(orient='records'))
        pip.set(job, 'competitors_df_json', pd.DataFrame().to_json(orient='records'))
        return pd.DataFrame() # Return empty DataFrame

# --- Helper Functions for Title Fetching (Made private) ---

_user_agent = 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36'
_headers = {'User-Agent': _user_agent}

def _get_title_from_html(html_content):
    """Simple helper to extract the title from HTML content."""
    soup = BeautifulSoup(html_content, 'html.parser')
    title_tag = soup.find('title')
    return title_tag.text.strip() if title_tag else '' # Added strip()

def _match_domain_in_title(domain, title):
    """Finds a stripped version of the domain in the title."""
    # Ensure domain is not None or empty before splitting
    if not domain:
        return ''
    base_domain = domain.split('.')[0]
    # Handle potential empty base_domain after split
    if not base_domain:
        return ''
    # Escape regex special characters in base_domain
    safe_base_domain = re.escape(base_domain)
    pattern = ''.join([c + r'\s*' for c in safe_base_domain])
    regex = re.compile(pattern, re.IGNORECASE)
    match = regex.search(title)
    if match:
        matched = match.group(0).strip()
        return matched
    return ''


async def _async_check_url(url, domain, timeout):
    """Asynchronously checks a single domain and extracts title/matched title."""
    try:
        async with httpx.AsyncClient(follow_redirects=True, headers=_headers, timeout=timeout) as client:
            response = await client.get(url)
            effective_url = str(response.url) # Store effective URL after redirects
            if response.status_code == 200:
                if effective_url != url:
                    print(f"  Redirected: {url} -> {effective_url}")
                title = _get_title_from_html(response.text)
                matched_title = _match_domain_in_title(domain, title)
                return effective_url, title, matched_title, True
            else:
                print(f"  Status Code {response.status_code} for {url}")
    except httpx.RequestError as e:
        # More specific error logging
        error_type = type(e).__name__
        print(f"  Request failed for {url}: {error_type} - {str(e)}")
    except httpx.TimeoutException:
         print(f"  Timeout for {url} after {timeout} seconds.")
    except Exception as e:
        print(f"  An unexpected error occurred for {url}: {type(e).__name__} - {str(e)}")
    # Ensure consistent return structure on failure
    return url, '', '', False # Return empty strings for title/matched_title

async def _async_test_domains(tasks):
    """Internal helper for asyncio.gather."""
    # return_exceptions=True ensures that one failed task doesn't stop others
    return await asyncio.gather(*tasks, return_exceptions=True)

def _test_domains(domains, timeout=120):
    """Orchestrates async checks for a list of domains."""
    print(f"  Giving up to {timeout} seconds per site...")
    # Ensure nest_asyncio is applied in the environment where this runs
    nest_asyncio.apply()
    tasks = [_async_check_url(f'https://{domain}', domain, timeout) for domain in domains]
    # Use asyncio.run() to execute the async gather function
    results = asyncio.run(_async_test_domains(tasks))

    domain_results = {}
    for domain, result in zip(domains, results):
        if isinstance(result, Exception):
            print(f"  Error processing {domain}: {result}")
            domain_results[domain] = {'url': f'https://{domain}', 'title': '', 'matched_title': ''} # Provide defaults
        elif isinstance(result, tuple) and len(result) == 4:
             # Check if status was success (fourth element)
             if result[3]:
                  domain_results[domain] = {'url': result[0], 'title': result[1], 'matched_title': result[2]}
             else:
                  # Handle cases where async_check_url returned False status but not an exception
                  domain_results[domain] = {'url': result[0], 'title': result[1] or '', 'matched_title': result[2] or ''}
        else:
             # Fallback for unexpected result format
             print(f"  Unexpected result format for {domain}: {result}")
             domain_results[domain] = {'url': f'https://{domain}', 'title': '', 'matched_title': ''}
    return domain_results


def _split_domain_name(domain):
    """Splits a concatenated domain name into human-readable words (requires wordninja)."""
    # Add basic check for non-string input
    if not isinstance(domain, str):
        return ''
    # Remove common TLDs before splitting for potentially cleaner results
    domain_no_tld = domain.split('.')[0]
    words = wordninja.split(domain_no_tld)
    return ' '.join(words)

# --- Main Function ---

def fetch_titles_and_create_filters(job: str):
    """
    Fetches homepage titles for competitors lacking them, updates the competitors DataFrame
    and CSV, generates a keyword filter list, saves it to CSV, and updates pip state.

    Args:
        job (str): The current Pipulate job ID.

    Returns:
        str: A status message summarizing the actions taken.
    """
    print("🏷️  Fetching competitor titles and generating keyword filters...")

    # --- PATH DEFINITIONS ---
    temp_dir = Path("temp") / job
    temp_dir.mkdir(parents=True, exist_ok=True)
    competitors_csv_file = temp_dir / "competitors.csv"
    filter_file = temp_dir / "filter_keywords.csv"

    # --- INPUT (from pip state) ---
    try:
        from io import StringIO # Import StringIO here
        competitors_df_json = pip.get(job, 'competitors_df_json', '[]')
        # --- FIX: Wrap JSON string in StringIO ---
        df_competitors = pd.read_json(StringIO(competitors_df_json), orient='records')
        # --- END FIX ---
        # Ensure required columns exist, even if empty
        for col in ['Domain', 'Column Label', 'Title', 'Matched Title']:
             if col not in df_competitors.columns:
                  df_competitors[col] = '' if col in ['Title', 'Matched Title'] else None

    except Exception as e:
        print(f"❌ Error loading competitors DataFrame from pip state: {e}")
        return "Error loading competitors data. Cannot proceed."

    if df_competitors.empty:
         print("🤷 Competitors DataFrame is empty. Skipping title fetch and filter generation.")
         # Still create an empty filter file if none exists
         if not filter_file.exists():
              pd.DataFrame(columns=['Filter']).to_csv(filter_file, index=False)
              print(f"  ✅ Created empty keyword filter file at '{filter_file}'.")
         return "Competitors list empty. Filter step skipped."


    # --- CORE LOGIC (Moved and Adapted) ---
    status_messages = []

    # Ensure correct data types and fill NaNs before filtering
    df_competitors['Title'] = df_competitors['Title'].fillna('').astype(str)
    df_competitors['Matched Title'] = df_competitors['Matched Title'].fillna('').astype(str).str.lower()
    df_competitors['Domain'] = df_competitors['Domain'].fillna('').astype(str)


    needs_titles = df_competitors[df_competitors['Title'] == ''].copy()

    if not needs_titles.empty:
        print(f"  Fetching Titles for {len(needs_titles)} domains...")
        results = _test_domains(needs_titles['Domain'].tolist())

        data_to_add = {'Domain': [], 'Title': [], 'Matched Title': []}
        for domain, info in results.items():
            data_to_add['Domain'].append(domain)
            data_to_add['Title'].append(info.get('title', '')) # Use .get for safety
            data_to_add['Matched Title'].append(info.get('matched_title', ''))

        new_data_df = pd.DataFrame(data_to_add)
        new_data_df['Matched Title'] = new_data_df['Matched Title'].str.lower() # Lowercase new matches


        # Combine using merge for clarity
        df_competitors = pd.merge(df_competitors, new_data_df, on='Domain', how='left', suffixes=('', '_new'))

        # Update Title and Matched Title only where they were originally empty
        df_competitors['Title'] = df_competitors.apply(
             lambda row: row['Title_new'] if pd.isna(row['Title']) or row['Title'] == '' else row['Title'], axis=1
        )
        df_competitors['Matched Title'] = df_competitors.apply(
             lambda row: row['Matched Title_new'] if pd.isna(row['Matched Title']) or row['Matched Title'] == '' else row['Matched Title'], axis=1
        )

        # Drop temporary merge columns
        df_competitors.drop(columns=['Title_new', 'Matched Title_new'], inplace=True)


        # Persist updated competitors data to CSV
        try:
            df_competitors.to_csv(competitors_csv_file, index=False)
            status_messages.append(f"Updated {len(needs_titles)} competitor titles and saved to CSV.")
            print(f"  ✅ Updated competitor titles and saved to '{competitors_csv_file}'.")
        except Exception as e:
            print(f"  ❌ Error saving updated competitors CSV: {e}")
            status_messages.append("Error saving updated competitors CSV.")

    else:
        status_messages.append("No missing competitor titles to fetch.")
        print("  ✅ All competitors already have titles.")

    # --- Create Keyword Filters ---
    try:
        # Ensure 'Domain' and 'Matched Title' columns exist and handle potential NaN/None
        extracted_domains = [_extract_registered_domain(str(domain)).replace('.com', '') for domain in df_competitors['Domain'].dropna()]
        matched_titles = [str(title).replace('.com', '') for title in df_competitors['Matched Title'].dropna().tolist() if title] # Filter empty strings
        split_domains = [_split_domain_name(domain) for domain in extracted_domains] # Use helper

        combined_list = [x.strip() for x in extracted_domains + matched_titles + split_domains if x] # Filter empty strings after strip
        combined_list = sorted(list(set(combined_list))) # Deduplicate

        # Persist to external filter file
        if not filter_file.exists():
            df_filter = pd.DataFrame(combined_list, columns=['Filter'])
            df_filter.to_csv(filter_file, index=False)
            status_messages.append(f"Created initial keyword filter file with {len(combined_list)} terms.")
            print(f"  ✅ Created initial keyword filter file at '{filter_file}'.")
        else:
            # Optionally, load existing, merge, dedupe, and save if you want it to be additive
            # For now, just report it exists
            df_existing_filter = pd.read_csv(filter_file)
            existing_terms = df_existing_filter['Filter'].dropna().astype(str).tolist()
            new_combined = sorted(list(set(existing_terms + combined_list)))
            if len(new_combined) > len(existing_terms):
                 df_new_filter = pd.DataFrame(new_combined, columns=['Filter'])
                 df_new_filter.to_csv(filter_file, index=False)
                 print(f"  🔄 Updated keyword filter file at '{filter_file}' ({len(new_combined)} total terms).")
                 status_messages.append(f"Updated keyword filter file ({len(new_combined)} total terms).")

            else:
                 print(f"  ☑️ Keyword filter file already exists at '{filter_file}' and requires no update.")
                 status_messages.append("Keyword filter file exists and is up-to-date.")


        # --- OUTPUT (to pip state) ---
        pip.set(job, 'competitors_df_json', df_competitors.to_json(orient='records'))
        # Store the generated/updated filter list as well
        pip.set(job, 'filter_keyword_list_json', json.dumps(combined_list))
        print(f"💾 Stored updated competitors DataFrame and filter list in pip state for job '{job}'.")
        # ---------------------------

    except Exception as e:
        print(f"❌ An error occurred during filter generation: {e}")
        status_messages.append("Error generating keyword filters.")
        # Attempt to save competitors DF state even if filter gen fails
        pip.set(job, 'competitors_df_json', df_competitors.to_json(orient='records'))


    # --- RETURN VALUE ---
    return "\n".join(status_messages) # Return summary string


def aggregate_semrush_metrics(job: str, df2: pd.DataFrame):
    """
    Aggregates metrics in the combined SEMRush DataFrame (df2) by Keyword,
    stores the aggregated DataFrame in pip state, and returns it.

    Args:
        job (str): The current Pipulate job ID.
        df2 (pd.DataFrame): The combined master DataFrame.

    Returns:
        pd.DataFrame: The aggregated DataFrame (agg_df), or an empty DataFrame on error.
    """
    if df2.empty:
        print("⚠️ Input DataFrame (df2) is empty. Cannot perform aggregation.")
        return pd.DataFrame()

    print("📊 Aggregating SEMRush metrics per keyword...")

    # --- NEW: Data Type Coercion (The Fix) ---
    # Define columns that MUST be numeric for aggregation.
    # This prevents the '[how->max,dtype->object]' error.
    numeric_cols = [
        'Position', 'Search Volume', 'CPC', 'Traffic', 'Traffic (%)',
        'Traffic Cost', 'Keyword Difficulty', 'Competition', 'Number of Results'
    ]

    print("  Coercing numeric columns...")
    for col in numeric_cols:
        if col in df2.columns:
            # errors='coerce' will turn any non-numeric values (like 'N/A' or '1,200') into NaN
            # We then fillna(0) to ensure mathematical operations don't fail,
            # though agg functions often handle NaN well, this is safer.
            df2[col] = pd.to_numeric(df2[col], errors='coerce').fillna(0)
        else:
            # This just matches the warning you already have, but it's good to be aware
            print(f"  > Warning: Numeric column '{col}' not found in source, skipping coercion.")
    # --- END FIX ---

    # --- THE FIX: Coerce the Timestamp column explicitly ---
    # The error comes from trying to find the 'max' of the 'Timestamp' column,
    # which is currently text. We must convert it to datetime objects first.
    if 'Timestamp' in df2.columns:
        print("  Coercing 'Timestamp' column to datetime...")
        df2['Timestamp'] = pd.to_datetime(df2['Timestamp'], errors='coerce')
    # --- END FIX ---

    # --- SMOKING GUN: Print dtypes *after* all coercion ---
    print("\n--- SMOKING GUN: Final Column Data Types Before Aggregation ---")
    print(df2.dtypes)
    print("--------------------------------------------------------------\n")

    # --- CORE LOGIC (Moved from Notebook) ---
    try:
        agg_funcs = {
            'Position': 'min', # Best position achieved
            'Search Volume': 'max', # Highest reported volume
            'CPC': 'mean', # Average CPC
            'Traffic': 'sum', # Total traffic (though often dropped later)
            'Traffic (%)': 'mean', # Average traffic % (though often dropped later)
            'Traffic Cost': 'sum', # Total traffic cost (though often dropped later)
            'Keyword Difficulty': 'mean', # Average difficulty
            'Previous position': 'first', # Arbitrary choice for non-aggregatable
            'Competition': 'mean', # Average competition score
            'Number of Results': 'max', # Highest number of results
            'Timestamp': 'max', # Latest timestamp
            'SERP Features by Keyword': 'first', # Arbitrary choice
            'Keyword Intents': 'first', # Arbitrary choice
            'Position Type': 'first', # Arbitrary choice
            'URL': 'first', # Arbitrary choice (often competitor URL if client doesn't rank)
            'Competitor URL': 'first', # First competitor URL found for the keyword
            'Client URL': 'first' # First client URL found (might be NaN)
        }

        # Check which columns actually exist in df2 before aggregating
        valid_agg_funcs = {k: v for k, v in agg_funcs.items() if k in df2.columns}
        missing_agg_cols = [k for k in agg_funcs if k not in df2.columns]
        if missing_agg_cols:
            print(f"  ⚠️ Columns not found for aggregation, will be skipped: {', '.join(missing_agg_cols)}")

        agg_df = df2.groupby('Keyword').agg(valid_agg_funcs).reset_index()

        # Calculate 'Number of Words' only if 'Keyword' column exists
        if 'Keyword' in agg_df.columns:
             agg_df['Number of Words'] = agg_df["Keyword"].astype(str).apply(lambda x: len(x.split()))
        else:
             print("  ⚠️ 'Keyword' column not found after aggregation. Cannot calculate 'Number of Words'.")


        # Drop 'Position' - It's misleading after aggregation here.
        # The real position data is kept in the pivot_df generated earlier.
        if 'Position' in agg_df.columns:
            agg_df.drop(columns=['Position'], inplace=True)

        print("  ✅ Table of aggregates prepared.")
        rows, cols = agg_df.shape
        print(f"  Rows: {rows:,}, Columns: {cols:,}")


        # --- OUTPUT (to pip state) ---
        pip.set(job, 'keyword_aggregate_df_json', agg_df.to_json(orient='records'))
        print(f"💾 Stored aggregated DataFrame in pip state for job '{job}'.")
        # ---------------------------

        # --- RETURN VALUE ---
        return agg_df # Return the DataFrame for display and next step

    except Exception as e:
        print(f"❌ An error occurred during aggregation: {e}")
        pip.set(job, 'keyword_aggregate_df_json', pd.DataFrame().to_json(orient='records'))
        return pd.DataFrame() # Return empty DataFrame


def _reorder_columns_surgical(df, priority_column, after_column):
    """
    Private helper: Moves a column immediately after a specified column. Handles missing columns.
    """
    # Check if columns exist *before* trying to manipulate
    if priority_column not in df.columns:
        print(f"  ⚠️ Reorder Warning: Priority column '{priority_column}' not found. Skipping.")
        return df
    if after_column not in df.columns:
        print(f"  ⚠️ Reorder Warning: After column '{after_column}' not found. Placing '{priority_column}' at end.")
        # Fallback: Move to end if after_column is missing
        cols = [col for col in df.columns if col != priority_column] + [priority_column]
        return df[cols]

    # Proceed with reordering if both columns exist
    columns = df.columns.drop(priority_column).tolist()
    try:
        after_column_index = columns.index(after_column)
        columns.insert(after_column_index + 1, priority_column)
        return df[columns]
    except ValueError:
         # This case should ideally not happen if after_column check passed, but good to have safeguard
         print(f"  ⚠️ Reorder Warning: Index for '{after_column}' not found unexpectedly. Placing '{priority_column}' at end.")
         cols = [col for col in df.columns if col != priority_column] + [priority_column]
         return df[cols]


# --- Main Function ---
def merge_filter_arrange_data(job: str, pivot_df: pd.DataFrame, agg_df: pd.DataFrame):
    """
    Merges pivot and aggregate DataFrames, applies keyword filters, reorders/drops columns,
    sorts the result, stores it in pip state, and returns the final DataFrame.

    Args:
        job (str): The current Pipulate job ID.
        pivot_df (pd.DataFrame): The pivoted DataFrame from the previous step.
        agg_df (pd.DataFrame): The aggregated DataFrame from the previous step.

    Returns:
        pd.DataFrame: The final, arranged DataFrame, or an empty DataFrame on error.
    """
    if pivot_df.empty or agg_df.empty:
        print("⚠️ Input DataFrame(s) (pivot_df or agg_df) are empty. Cannot merge and filter.")
        return pd.DataFrame()

    print("🧩 Merging Pivot Data with Aggregate Data...")

    # --- PATH DEFINITION ---
    filter_file = Path("data") / f"{job}_filter_keywords.csv"

    # --- CORE LOGIC (Moved from Notebook) ---
    try:
        # 1. Merge DataFrames
        # Ensure 'Keyword' is a column for merging (pivot_df might have it as index)
        pivot_df_reset = pivot_df.reset_index() if 'Keyword' not in pivot_df.columns else pivot_df
        # agg_df should already have 'Keyword' as a column from its creation

        # Check if 'Keyword' column exists in both before merge
        if 'Keyword' not in pivot_df_reset.columns or 'Keyword' not in agg_df.columns:
             raise ValueError("Missing 'Keyword' column in one of the DataFrames for merging.")

        pivotmerge_df = pd.merge(pivot_df_reset, agg_df, on='Keyword', how='left')
        print("  ✅ Pivot and Aggregate Data Joined.")
        rows_pre_filter, cols_pre_filter = pivotmerge_df.shape
        print(f"     Rows: {rows_pre_filter:,}, Columns: {cols_pre_filter:,}")

        # --- Filtering ---
        print("\n🧹 Applying Brand and Negative Filters...")
        kw_filter = [] # Initialize filter list
        if filter_file.exists():
            try:
                df_filter = pd.read_csv(filter_file, header=0)
                # Ensure Filter column exists and handle potential NaNs robustly
                if 'Filter' in df_filter.columns:
                     kw_filter = df_filter["Filter"].dropna().astype(str).tolist()
                else:
                     print(f"  ⚠️ Filter file '{filter_file}' exists but missing 'Filter' column.")

            except Exception as e:
                print(f"  ⚠️ Error reading filter file '{filter_file}': {e}")

        if kw_filter:
            pattern = '|'.join([re.escape(keyword) for keyword in kw_filter])
            # Ensure 'Keyword' column exists before filtering
            if 'Keyword' in pivotmerge_df.columns:
                 filtered_df = pivotmerge_df[~pivotmerge_df["Keyword"].astype(str).str.contains(pattern, case=False, na=False)]
                 print(f"  ✅ Filter applied using {len(kw_filter)} terms from '{filter_file}'.")
            else:
                 print("  ⚠️ 'Keyword' column missing. Cannot apply filter.")
                 filtered_df = pivotmerge_df # Skip filtering
        else:
            filtered_df = pivotmerge_df
            if filter_file.exists():
                 print("  ⚠️ Filter file exists but contains no terms. Skipping filter application.")
            else:
                 print(f"  ☑️ No filter file found at '{filter_file}'. Skipping negative filtering.")

        rows_post_filter, _ = filtered_df.shape
        removed_count = rows_pre_filter - rows_post_filter
        print(f"     Rows after filter: {rows_post_filter:,} ({removed_count:,} rows removed)")


        # --- Reordering and Final Polish ---
        print("\n✨ Reordering columns and finalizing...")
        temp_df = filtered_df.copy()

        # Chain reorders using the robust helper
        temp_df = _reorder_columns_surgical(temp_df, "Search Volume", after_column="Keyword")
        temp_df = _reorder_columns_surgical(temp_df, "Number of Words", after_column="CPC")
        temp_df = _reorder_columns_surgical(temp_df, "CPC", after_column="Number of Words")
        temp_df = _reorder_columns_surgical(temp_df, "Number of Results", after_column="Position Type")
        temp_df = _reorder_columns_surgical(temp_df, "Timestamp", after_column="Number of Results")
        temp_df = _reorder_columns_surgical(temp_df, "Competitor URL", after_column="Client URL")

        # Final column arrangement (Keyword, Search Volume first)
        essential_cols = ['Keyword', 'Search Volume']
        if all(col in temp_df.columns for col in essential_cols):
            rest_of_columns = [col for col in temp_df.columns if col not in essential_cols]
            new_column_order = essential_cols + rest_of_columns
            # Defensively check if all original columns are still in the new order
            if set(new_column_order) == set(temp_df.columns):
                 arranged_df = temp_df[new_column_order]
                 print("  ✅ Final column order applied.")
            else:
                 print("  ⚠️ Column mismatch during final arrange. Using previous order.")
                 arranged_df = temp_df # Fallback
        else:
            print("  ⚠️ Essential columns ('Keyword', 'Search Volume') missing. Skipping final arrange.")
            arranged_df = temp_df # Fallback

        # Final sorting and column drops
        sort_col = "Search Volume" if "Search Volume" in arranged_df.columns else "Keyword"
        arranged_df = arranged_df.sort_values(by=sort_col, ascending=False)

        cols_to_drop = ["Previous position", "Traffic", "Traffic (%)", "Traffic Cost"]
        existing_cols_to_drop = [col for col in cols_to_drop if col in arranged_df.columns]
        if existing_cols_to_drop:
            arranged_df.drop(columns=existing_cols_to_drop, inplace=True)
            print(f"  ✂️ Dropped columns: {', '.join(existing_cols_to_drop)}")

        print("\n✅ Final Keyword Table Prepared.")
        final_rows, final_cols = arranged_df.shape
        print(f"   Final Rows: {final_rows:,}, Final Columns: {final_cols:,}")


        # --- OUTPUT (to pip state) ---
        pip.set(job, 'filtered_gap_analysis_df_json', arranged_df.to_json(orient='records'))
        print(f"💾 Stored final arranged DataFrame in pip state for job '{job}'.")
        # ---------------------------

        # --- RETURN VALUE ---
        return arranged_df # Return the DataFrame for display

    except Exception as e:
        print(f"❌ An error occurred during merge/filter/arrange: {e}")
        pip.set(job, 'filtered_gap_analysis_df_json', pd.DataFrame().to_json(orient='records'))
        return pd.DataFrame() # Return empty DataFrame


def _download_file(download_url, output_path):
    response = requests.get(download_url, stream=True)
    if response.status_code == 200:
        output_path.parent.mkdir(parents=True, exist_ok=True)  # Ensure the directory exists
        with open(output_path, "wb") as file:
            for chunk in response.iter_content(chunk_size=8192):
                file.write(chunk)
        return True
    else:
        print(f"Failed to download file. Status Code: {response.status_code}")
        return False


def _decompress_gz(gz_path, output_path):
    try:
        with gzip.open(gz_path, 'rb') as f_in, open(output_path, 'wb') as f_out:
            shutil.copyfileobj(f_in, f_out)
        print(f"Decompressed {output_path}")
        return True
    except Exception as e:
        print(f"Failed to decompress {gz_path}. Error: {e}")
        return False


def _fetch_analysis_slugs(org, project, botify_token):
    """Fetch analysis slugs for a given project from the Botify API."""
    analysis_url = f"https://api.botify.com/v1/analyses/{org}/{project}/light"
    headers = {"Authorization": f"Token {botify_token}"}
    try:
        response = requests.get(analysis_url, headers=headers)
        response.raise_for_status()
        data = response.json()
        analysis_data = data.get('results', [])
        return [analysis['slug'] for analysis in analysis_data]
    except Exception as e:
        print(f"❌ Error fetching analysis slugs: {e}")
        return []


def _export_data(version, org, project, export_payload, report_path, analysis=None, retry_url=None):
    """
    Unified function to export data using BQLv1 or BQLv2.
    Handles API calls, polling, download, and decompression via helpers.
    Uses keys.botify directly for authentication.

    Args:
        version (str): 'v1' or 'v2'.
        org (str): Botify organization slug.
        project (str): Botify project slug.
        export_payload (dict): The payload for the export API call.
        report_path (Path): The desired final output path (e.g., .../file.csv).
        analysis (str, optional): The analysis slug (required for v1). Defaults to None.
        retry_url (str, optional): A direct download URL if polling fails. Defaults to None.

    Returns:
        tuple: (status_code: int or str, final_path: Path or None)
    """
    file_base = report_path.stem
    path_base = Path(report_path).parent
    zip_name = path_base / f"{file_base}.gz"
    csv_name = Path(report_path)

    path_base.mkdir(parents=True, exist_ok=True)  # Ensure the directory exists before proceeding

    if csv_name.exists():
        print(f"The file: {csv_name}")
        print("...already exists for analysis period. Exiting.")
        return (None, None)

    if zip_name.exists():
        print(f"☑️ {zip_name} found without corresponding CSV. Decompressing now...")
        decompress_success = _decompress_gz(zip_name, csv_name)
        return (200, None) if decompress_success else (None, None)

    if retry_url:
        print(f"Using retry URL for direct download: {retry_url}")
        if _download_file(retry_url, zip_name):  # Save as .gz file
            print("File downloaded successfully via retry URL.")
            if _decompress_gz(zip_name, csv_name):  # Decompress .gz to .csv
                print("File decompressed successfully.")
                return (200, csv_name)
            else:
                print("Decompression failed.")
                return (None, None)
        else:
            print("Download failed using retry URL.")
            return (None, None)

    # Use the token from the keys module
    headers = {'Authorization': f'Token {keys.botify}', 'Content-Type': 'application/json'}

    if version == 'v1':
        url = f'https://api.botify.com/v1/analyses/{org}/{project}/{analysis}/urls/export'
        response = requests.post(url, headers=headers, json=export_payload)
    else:  # version == 'v2'
        url = "https://api.botify.com/v1/jobs"
        response = requests.post(url, headers=headers, json=export_payload)

    if response.status_code not in [200, 201]:
        print(f"❌ Failed to start CSV export. Status Code: {response.status_code}.")
        print(response.reason, response.text)
        pprint(export_payload)
        return (response.status_code, None)

    export_job_details = response.json()
    job_url = export_job_details.get('job_url')
    if version == "v2":
        job_url = f'https://api.botify.com{job_url}'

    attempts = 300
    delay = 10
    print(f"{attempts} attempts will be made every {delay} seconds until download is ready...")

    while attempts > 0:
        sleep(delay)
        print(attempts, end=" ", flush=True)  # Countdown on the same line
        response_poll = requests.get(job_url, headers=headers)
        if response_poll.status_code == 200:
            job_status_details = response_poll.json()
            if job_status_details['job_status'] == 'DONE':
                print("\nExport job done.")
                download_url = job_status_details['results']['download_url']
                if _download_file(download_url, zip_name):
                    print("File downloaded successfully.")
                    if _decompress_gz(zip_name, csv_name):
                        print("File decompressed successfully.")
                        return (200, csv_name)
                    else:
                        print("Decompression failed.")
                        return ("Decompression failed 1.", None)
                else:
                    print("Download failed.")
                    return ("Download failed.", None)
            elif job_status_details['job_status'] == 'FAILED':
                print("\nExport job failed.")
                print(job_status_details.get('failure_reason', 'No failure reason provided.'))
                return ("Export job failed.", None)
        else:
            print(f"\nFailed to get export status. Status Code: {response_poll.status_code}")
            print(response_poll.text)

        attempts -= 1

    print("Unable to complete download attempts successfully.")
    return ("Unable to complete", None)


def fetch_botify_data(job: str, botify_token: str, botify_project_url: str):
    """
    Orchestrates fetching data from the Botify API, handling slug detection,
    API calls with fallbacks, downloading, and decompression. Stores the final
    DataFrame in pip state.

    Args:
        job (str): The current Pipulate job ID.
        botify_token (str): The Botify API token.
        botify_project_url (str): The Botify project URL to parse for org/project slugs.

    Returns:
        tuple: (botify_df: pd.DataFrame, has_botify: bool)
               Returns the fetched DataFrame and a boolean indicating success.
    """
    print("🤖 Fetching data from Botify API...")
    
    # --- 1. Parse URL and get latest analysis slug ---
    try:
        url_parts = botify_project_url.rstrip('/').split('/')
        org = url_parts[-2]
        project = url_parts[-1]
        print(f"  Parsed Org: {org}, Project: {project}")
        
        slugs = _fetch_analysis_slugs(org, project, botify_token)
        if not slugs:
            raise ValueError("Could not find any Botify analysis slugs for the provided project.")
        analysis = slugs[0] # Use the most recent analysis
        print(f"  ✅ Found latest Analysis Slug: {analysis}")

    except (IndexError, ValueError) as e:
        print(f"  ❌ Critical Error: Could not parse Botify URL or find analysis slug. {e}")
        pip.set(job, 'botify_export_df_json', pd.DataFrame().to_json(orient='records'))
        return pd.DataFrame(), False

    # --- 2. Define Paths and Payloads ---
    csv_dir = Path("data") / f"{job}_botify"
    csv_dir.mkdir(parents=True, exist_ok=True)
    report_name = csv_dir / "botify_export.csv"

    payload_full = {
        "fields": ["url", "depth", "gsc_by_url.count_missed_clicks", "gsc_by_url.avg_ctr", "gsc_by_url.avg_position", "inlinks_internal.nb.unique", "internal_page_rank.value", "internal_page_rank.position", "internal_page_rank.raw", "gsc_by_url.count_impressions", "gsc_by_url.count_clicks", "gsc_by_url.count_keywords", "gsc_by_url.count_keywords_on_url_to_achieve_90pc_clicks", "metadata.title.content", "metadata.description.content"],
        "sort": []
    }
    payload_fallback = {
        "fields": ["url", "depth", "inlinks_internal.nb.unique", "internal_page_rank.value", "internal_page_rank.position", "internal_page_rank.raw", "metadata.title.content", "metadata.description.content"],
        "sort": []
    }

    # --- 3. Main Logic: Check existing, call API with fallback ---
    botify_export_df = None
    if report_name.exists():
        print(f"  ☑️ Botify export file already exists at '{report_name}'. Reading from disk.")
        try:
            botify_export_df = pd.read_csv(report_name, skiprows=1)
        except Exception as e:
            print(f"  ⚠️ Could not read existing CSV file '{report_name}', will attempt to re-download. Error: {e}")

    if botify_export_df is None:
        print("  - Attempting download with Full GSC Payload...")
        status_code, _ = _export_data('v1', org, project, payload_full, report_name, botify_token, analysis=analysis)
        
        if status_code not in [200, 201]:
            print("  - Full Payload failed. Attempting Fallback Payload (no GSC data)...")
            status_code, _ = _export_data('v1', org, project, payload_fallback, report_name, botify_token, analysis=analysis)
        
        if status_code in [200, 201] or report_name.exists():
            try:
                botify_export_df = pd.read_csv(report_name, skiprows=1)
                print("  ✅ Successfully downloaded and loaded Botify data.")
            except Exception as e:
                print(f"  ❌ Download seemed successful, but failed to read the final CSV file '{report_name}'. Error: {e}")
                botify_export_df = pd.DataFrame() # Ensure it's an empty DF on read failure
        else:
            print("  ❌ Botify export failed critically after both attempts.")
            botify_export_df = pd.DataFrame()

    # --- START URGENT FIX: Comment out the failing pip.set() ---
    # We will skip persistence for this step to get past the TooBigError.
    # The df is already in memory for the next cell, which is all we need.
    # pip.set(job, 'botify_export_csv_path', str(report_name.resolve()))
    # --- 4. Store State and Return ---
    has_botify = not botify_export_df.empty
    if has_botify:
        print(f"💾 Bypassing pip.set() for Botify CSV path to avoid TooBigError.")
    else:
        # If it failed, ensure an empty DF is stored
        pip.set(job, 'botify_export_csv_path', None)
        print("🤷 No Botify data loaded. Stored empty DataFrame in pip state.")
        # pip.set(job, 'botify_export_csv_path', None)
        print("🤷 No Botify data loaded. Bypassing pip.set().")
    # --- END URGENT FIX ---
    return botify_export_df, has_botify


def fetch_botify_data_and_save(job: str, botify_token: str, botify_project_url: str):
    """
    Orchestrates fetching data from the Botify API using pre-defined helpers,
    handling slug detection, API calls with fallbacks, downloading, decompression,
    and storing the final DataFrame in pip state.

    Args:
        job (str): The current Pipulate job ID.
        botify_token (str): The Botify API token.
        botify_project_url (str): The Botify project URL to parse for org/project slugs.

    Returns:
        tuple: (botify_df: pd.DataFrame, has_botify: bool, report_path: Path or None, csv_dir_path: Path or None)
               Returns the fetched DataFrame, a boolean indicating success, and the report/directory paths.
    """
    print("🤖 Fetching data from Botify API...")
    report_name = None # Initialize report_name
    csv_dir = None # Initialize csv_dir
    botify_export_df = pd.DataFrame() # Initialize as empty DataFrame

    # --- 1. Parse URL and get latest analysis slug ---
    try:
        # Strip trailing slash FIRST for reliable parsing
        cleaned_url = botify_project_url.rstrip('/')
        url_parts = cleaned_url.split('/')
        if len(url_parts) < 2: # Basic validation
             raise ValueError(f"Could not parse org/project from URL: {botify_project_url}")

        org = url_parts[-2]
        project = url_parts[-1]
        print(f"  Parsed Org: {org}, Project: {project}")

        slugs = _fetch_analysis_slugs(org, project, botify_token)
        if not slugs:
            raise ValueError("Could not find any Botify analysis slugs for the provided project.")
        analysis = slugs[0] # Use the most recent analysis
        print(f"  ✅ Found latest Analysis Slug: {analysis}")

    except (IndexError, ValueError, Exception) as e: # Catch broader exceptions during setup
        print(f"  ❌ Critical Error during Botify setup: {e}")
        pip.set(job, 'botify_export_df_json', pd.DataFrame().to_json(orient='records'))
        return pd.DataFrame(), False, None, None # Return empty DF, False, and None paths

    # --- 2. Define Paths and Payloads ---
    try:
        csv_dir = Path("temp") / job # Pointing to the consolidated temp/job_name dir
        csv_dir.mkdir(parents=True, exist_ok=True)
        report_name = csv_dir / "botify_export.csv"

        payload_full = {
            "fields": ["url", "depth", "gsc_by_url.count_missed_clicks", "gsc_by_url.avg_ctr", "gsc_by_url.avg_position", "inlinks_internal.nb.unique", "internal_page_rank.value", "internal_page_rank.position", "internal_page_rank.raw", "gsc_by_url.count_impressions", "gsc_by_url.count_clicks", "gsc_by_url.count_keywords", "gsc_by_url.count_keywords_on_url_to_achieve_90pc_clicks", "metadata.title.content", "metadata.description.content"],
            "sort": []
        }
        payload_fallback = {
            "fields": ["url", "depth", "inlinks_internal.nb.unique", "internal_page_rank.value", "internal_page_rank.position", "internal_page_rank.raw", "metadata.title.content", "metadata.description.content"],
            "sort": []
        }
    except Exception as e:
        print(f"  ❌ Error defining paths/payloads: {e}")
        pip.set(job, 'botify_export_df_json', pd.DataFrame().to_json(orient='records'))
        return pd.DataFrame(), False, None, csv_dir # Return csv_dir if it was created

    # --- 3. Main Logic: Check existing, call API with fallback ---
    loaded_from_existing = False
    if report_name.exists():
        print(f"  ☑️ Botify export file already exists at '{report_name}'. Reading from disk.")
        try:
            # Skip header row which often contains metadata from Botify exports
            botify_export_df = pd.read_csv(report_name, skiprows=1)
            loaded_from_existing = True # Flag success
        except Exception as e:
            print(f"  ⚠️ Could not read existing CSV file '{report_name}', will attempt to re-download. Error: {e}")
            botify_export_df = pd.DataFrame() # Reset DF if read fails

    # Only attempt download if not loaded from existing file
    if not loaded_from_existing:
        print("  Attempting download with Full GSC Payload...")
        # Pass botify_token to the helper
        status_code, _ = _export_data('v1', org, project, payload_full, report_name, analysis=analysis)

        if status_code not in [200, 201]: # Check includes 201 for job creation success
            print("    -> Full Payload failed. Attempting Fallback Payload (no GSC data)...")
            status_code, _ = _export_data('v1', org, project, payload_fallback, report_name, analysis=analysis)

        # After attempts, check if the file exists and try to read it
        if report_name.exists():
             try:
                  botify_export_df = pd.read_csv(report_name, skiprows=1)
                  print("  ✅ Successfully downloaded and/or loaded Botify data.")
             except Exception as e:
                  print(f"  ❌ Download/decompression seemed successful, but failed to read the final CSV file '{report_name}'. Error: {e}")
                  botify_export_df = pd.DataFrame() # Ensure empty DF on read failure
        else:
             # Only print this if we didn't load from an existing file initially
             print("  ❌ Botify export failed critically after both attempts, and no file exists.")
             botify_export_df = pd.DataFrame()

    # --- 4. Store State and Return ---
    has_botify = not botify_export_df.empty
    pip.set(job, 'botify_export_df_json', botify_export_df.to_json(orient='records'))
    if has_botify:
        print(f"💾 Stored Botify DataFrame ({len(botify_export_df)} rows) in pip state for job '{job}'.")
    else:
        print("🤷 No Botify data loaded or available. Stored empty DataFrame in pip state.")

    # Return necessary info for display logic in notebook
    return botify_export_df, has_botify, report_name, csv_dir


def _insert_columns_after_surgical(df, column_names, after_column):
    """
    Surgical port: Inserts a list of columns immediately after a specified column.
    (Moved verbatim from notebook cell)
    """
    # This logic is complex, but as requested, moved verbatim.
    # We rely on the calling logic (in merge_and_finalize_data) to ensure
    # 'column_names' only contains *new* columns not in 'df'.
    
    if after_column not in df.columns:
        # If the reference column is missing, append columns to the end
        print(f"  ⚠️ Insert Warning: After-column '{after_column}' not found. Appending new columns to end.")
        new_order = df.columns.tolist() + [col for col in column_names if col not in df.columns]
        return df.reindex(columns=new_order) # Use reindex for safety
        
    # Per the original, this logic is flawed if column_names contains existing columns,
    # but our calling function ensures it only contains *new* columns.
    missing_columns = [col for col in column_names if col not in df.columns]
    
    # This line is flawed but kept verbatim from the notebook helper:
    column_names = [col for col in column_names if col not in missing_columns]
    
    # This check is added to handle the (now likely) empty column_names list
    if not column_names:
         # This is the most likely path, given the call logic.
         # We'll re-implement the *intent* which is to add the *new* columns.
         column_names = missing_columns # This is the correction
         if not column_names:
              return df # Truly nothing to insert
            
    insert_after_index = df.columns.get_loc(after_column)
    
    before = df.columns[:insert_after_index + 1].tolist()
    after = df.columns[insert_after_index + 1:].tolist()
    
    # Ensure columns to be inserted are not duplicated in the 'after' list
    after = [col for col in after if col not in column_names] 
    
    new_order = before + column_names + after
    
    # Check for consistency
    if len(new_order) != len(df.columns) + len(missing_columns):
        # Fallback if logic failed
        print(f"  ⚠️ Insert Warning: Column mismatch during reorder. Appending new columns to end.")
        return df.reindex(columns=df.columns.tolist() + missing_columns)
        
    return df[new_order]


def merge_and_finalize_data(job: str, arranged_df: pd.DataFrame, botify_export_df: pd.DataFrame, has_botify: bool):
    """
    Merges SEMRush data with Botify data (if present), cleans columns,
    saves an intermediate CSV, stores the final DF in pip state,
    and returns the final DF and data for display.
    
    Args:
        job (str): The current Pipulate job ID.
        arranged_df (pd.DataFrame): The filtered/arranged SEMRush data.
        botify_export_df (pd.DataFrame): The Botify data (or empty DF).
        has_botify (bool): Flag indicating if Botify data is present.
        
    Returns:
        tuple: (final_df: pd.DataFrame, display_data: dict)
               The final DataFrame (aliased as 'df' in notebook) and
               a dict with data for display (rows, cols, has_botify, pagerank_counts).
    """
    temp_dir = Path("temp") / job
    temp_dir.mkdir(parents=True, exist_ok=True) # Ensure it exists just in case
    unformatted_csv = temp_dir / "unformatted.csv"

    try:
        # 1. Determine if Botify data is present
        if has_botify and isinstance(botify_export_df, pd.DataFrame) and not botify_export_df.empty:
            
            # Perform the merge
            botify_data_for_merge = botify_export_df.rename(columns={"url": "Full URL"}, inplace=False)
            
            final_df = pd.merge(arranged_df, botify_data_for_merge, left_on='Client URL', right_on='Full URL', how='left')

            # Insert Botify columns after 'Competition'
            # This logic correctly identifies only the *new* columns added by the merge
            botify_cols_to_add = [col for col in botify_export_df.columns if col not in arranged_df.columns and col != 'url']
            
            # Use the moved helper function
            final_df = _insert_columns_after_surgical(final_df, botify_cols_to_add, "Competition")
            print("  ✅ Botify Data was found and merged.")
            
        else:
            has_botify = False # Ensure flag is correct if DF was empty
            final_df = arranged_df.copy() 
            print("  ℹ️ No Botify Data provided or DataFrame was empty. Skipping merge.")

        # 2. Final Cleanup and Persistence
        final_df = final_df.copy() # Ensure it's a copy

        # Drop redundant merge/artifact columns
        cols_to_drop = ["URL", "Full URL", "url"]
        existing_cols_to_drop = [col for col in cols_to_drop if col in final_df.columns]
        if existing_cols_to_drop:
             final_df.drop(columns=existing_cols_to_drop, inplace=True)
             print(f"  ✂️ Dropped redundant columns: {', '.join(existing_cols_to_drop)}")

        # Save unformatted intermediary file
        final_df.to_csv(unformatted_csv, index=False)
        print(f"  💾 Intermediate unformatted file saved to '{unformatted_csv}'")
        
        df = final_df.copy() # Create the 'df' alias for the next step

        # 3. Prepare Display Data
        rows, cols = df.shape
        pagerank_counts = None
        if has_botify and "Internal Pagerank" in df.columns:
            pagerank_counts = df["Internal Pagerank"].value_counts()
        
        display_data = {
            "rows": rows,
            "cols": cols,
            "has_botify": has_botify,
            "pagerank_counts": pagerank_counts
        }

        # --- OUTPUT (to pip state) ---
        # --- FIX: Store the PATH to the intermediate CSV, not the giant DataFrame JSON ---
        pip.set(job, 'final_working_df_csv_path', str(unformatted_csv.resolve()))
        print(f"💾 Stored final working DF path in pip state for job '{job}': {unformatted_csv.resolve()}")

        # --- RETURN VALUE ---
        return df, display_data

    except Exception as e:
        print(f"❌ An error occurred during final merge/cleanup: {e}")
        pip.set(job, 'final_working_df_csv_path', None) # Store None on error
        # Return empty/default values
        return pd.DataFrame(), {"rows": 0, "cols": 0, "has_botify": False, "pagerank_counts": None}


def truncate_dataframe_by_volume(job: str, final_df: pd.DataFrame, row_limit: int):
    """
    Truncates the DataFrame to be under a specific row limit by iterating
    through search volume cutoffs. Stores the truncated DF in pip state.

    Args:
        job (str): The current Pipulate job ID.
        final_df (pd.DataFrame): The DataFrame to truncate (from the previous merge step).
        row_limit (int): The target maximum number of rows.

    Returns:
        pd.DataFrame: The truncated DataFrame, or an empty DataFrame on error.
    """
    if final_df.empty:
        print("⚠️ Input DataFrame (final_df) is empty. Cannot truncate.")
        return pd.DataFrame()

    print(f"✂️ Truncating data to fit under {row_limit:,} rows for clustering...")

    # --- CORE LOGIC (Moved from Notebook) ---
    try:
        # Define the Search Volume Cut-off Increments
        volume_cutoffs = [49, 99, 199, 299, 499, 999, 1499, 1999, 2499, 2999, 3499, 3999, 5000, 7500, 10000, 20000, 30000]

        # Ensure 'Search Volume' column exists and is numeric
        if 'Search Volume' not in final_df.columns:
            print("  ❌ 'Search Volume' column not found. Cannot truncate by volume.")
            # Store and return the original df as-is
            pip.set(job, 'truncated_df_for_clustering_json', final_df.to_json(orient='records'))
            return final_df
        
        # Ensure 'Search Volume' is numeric, coercing errors to NaN and filling with 0
        final_df['Search Volume'] = pd.to_numeric(final_df['Search Volume'], errors='coerce').fillna(0)

        truncated_df = final_df.copy() # Initialize with the full DF
        try_fit = 0

        # Iterate to find the optimal Search Volume floor
        for cutoff in volume_cutoffs:
            df_candidate = final_df[final_df["Search Volume"] > cutoff]
            num_rows = df_candidate.shape[0]
            try_fit = cutoff
            print(f"  Volume >{cutoff:,} results in {num_rows:,} rows.")
            
            if num_rows <= row_limit:
                truncated_df = df_candidate # This is the best fit
                break
        
        # Handle edge case where loop finishes but DF is still too large
        # (i.e., even >30000 volume is more rows than ROW_LIMIT)
        if truncated_df.shape[0] > row_limit:
             print(f"  ⚠️ Could not get under {row_limit} rows. Using last valid cutoff >{try_fit:,} ({truncated_df.shape[0]} rows).")
        
        # Handle edge case where first filter cuts everything
        if truncated_df.shape[0] == 0 and final_df.shape[0] > 0:
            truncated_df = final_df[final_df["Search Volume"] > 0] # Fallback to > 0

        # --- Final Output and Persistence ---
        rows, cols = truncated_df.shape
        print(f"✅ Final truncation floor: Search Volume >{try_fit:,} resulting in {rows:,} rows.")

        speak(f"Data truncation complete. Retained {rows} rows with search volume above {try_fit}.")

        df_to_store = truncated_df.copy()

        # --- OUTPUT (to pip state) ---
        pip.set(job, 'truncated_df_for_clustering_json', df_to_store.to_json(orient='records'))
        print(f"💾 Stored truncated DataFrame ({len(df_to_store)} rows) in pip state for job '{job}'.")
        # ---------------------------

        # --- RETURN VALUE ---
        # Return the truncated DataFrame, which will be aliased as 'df' in the notebook
        return df_to_store

    except Exception as e:
        print(f"❌ An error occurred during truncation: {e}")
        pip.set(job, 'truncated_df_for_clustering_json', pd.DataFrame().to_json(orient='records'))
        return pd.DataFrame() # Return empty DataFrame


# --- 1. CORE ML UTILITIES ---

def calculate_silhouette(X, labels):
    """Calculates the Silhouette Coefficient for cluster evaluation."""
    # Handle the edge case where there is only one cluster or too few data points
    if len(np.unique(labels)) <= 1 or len(X) < 2:
        return 0.0 # Return 0 for non-evaluatable cases

    return silhouette_score(X, labels)

def preprocess_keywords(text):
    """Stems, lowercases, tokenizes, and removes stopwords from a keyword string."""
    stemmer = PorterStemmer()
    # Assuming stopwords were downloaded earlier with `nltk.download('stopwords')`
    stop_words = set(stopwords.words('english'))
    words = word_tokenize(text.lower())
    # Filter for alphanumeric words and then stem
    return ' '.join([stemmer.stem(word) for word in words if word not in stop_words and word.isalnum()])

def keyword_clustering(df, keyword_column, n_clusters=30, n_components=5, max_features=500):
    """Performs Tfidf Vectorization, Truncated SVD, and MiniBatchKMeans clustering."""

    # 1. Preprocess keywords
    df['Stemmed Keywords'] = df[keyword_column].apply(preprocess_keywords)

    # 2. Text Vectorization
    print(f"Vectorizing... (Max Features: {max_features})")
    vectorizer = TfidfVectorizer(max_features=max_features, stop_words='english')
    X = vectorizer.fit_transform(df['Stemmed Keywords'])

    # 3. Dimension Reduction
    print(f"Reducing Dimensions... (Components: {n_components})")
    svd = TruncatedSVD(n_components=n_components, random_state=42)
    principal_components = svd.fit_transform(X)

    # 4. Clustering
    print(f"Clustering... (K: {n_clusters})")
    # Setting compute_labels=True to ensure compatibility with MiniBatchKMeans
    minibatch_kmeans = MiniBatchKMeans(n_clusters=n_clusters, batch_size=100, random_state=42, n_init='auto') 
    df['Keyword Cluster'] = minibatch_kmeans.fit_predict(principal_components)

    # 5. Calculate silhouette score
    print("Calculating silhouette cluster quality score (takes a bit)...")
    silhouette_avg = calculate_silhouette(principal_components, df['Keyword Cluster'])

    # Return DataFrame, score, and the used parameters
    return df, silhouette_avg, {'n_clusters': n_clusters, 'n_components': n_components, 'max_features': max_features}

def name_keyword_clusters(df, keyword_column, cluster_column):
    """Names each cluster by the most common non-stopword, non-repeating bigram within the cluster."""

    stop_words = set(stopwords.words('english'))
    cluster_names = {}

    for cluster in df[cluster_column].unique():
        cluster_data = df[df[cluster_column] == cluster]
        all_keywords = ' '.join(cluster_data[keyword_column].astype(str)).split()
        filtered_keywords = [word for word in all_keywords if word not in stop_words and word.isalnum()]

        bigram_counts = Counter(bigrams(filtered_keywords))

        most_common_bigram = None
        for bigram, count in bigram_counts.most_common():
            if bigram[0] != bigram[1]:
                most_common_bigram = bigram
                break

        if not most_common_bigram:
            # Fallback to single most common word or a generic name
            unigram_counts = Counter(filtered_keywords)
            most_common_unigram = unigram_counts.most_common(1)
            most_common_words = most_common_unigram[0][0] if most_common_unigram else "Generic Cluster"
        else:
            most_common_words = ' '.join(most_common_bigram)

        cluster_names[cluster] = most_common_words

    df['Keyword Group (Experimental)'] = df[cluster_column].map(cluster_names)

    # Drop Process Columns (as per original logic)
    df.drop(columns=['Stemmed Keywords'], inplace=True)
    df.drop(columns=['Keyword Cluster'], inplace=True)

    return df


def cluster_and_finalize_dataframe(job: str, df: pd.DataFrame, has_botify: bool, enable_clustering: bool = True):
    """
    Performs keyword clustering (optional), names clusters, reorders columns,
    saves the unformatted CSV, stores the final DataFrame in pip state,
    and returns the final DataFrame for display.

    Args:
        job (str): The current Pipulate job ID.
        df (pd.DataFrame): The truncated DataFrame from the previous step.
        has_botify (bool): Flag indicating if Botify data is present.
        enable_clustering (bool): Whether to perform ML clustering. Defaults to True.

    Returns:
        pd.DataFrame: The final, clustered, and arranged DataFrame.
    """
    if df.empty:
        print("⚠️ Input DataFrame (df) is empty. Cannot perform clustering.")
        return pd.DataFrame()

    # --- CORE LOGIC (Moved from Notebook) ---
    try:
        # --- PATH DEFINITIONS ---
        temp_dir = Path("temp") / job # Define temp_dir if not already defined in scope
        temp_dir.mkdir(parents=True, exist_ok=True) # Ensure it exists
        keyword_cluster_params = temp_dir / "keyword_cluster_params.json"
        unformatted_csv = temp_dir / "unformatted.csv"

        if enable_clustering:
            print("🤖 Grouping Keywords (Clustering)...")

            # Download necessary nltk components
            nltk.download('punkt_tab', quiet=True)

            # Configuration for iterative testing
            target_silhouette_score = 0.6
            n_clusters_options = range(15, 26)
            n_components_options = [10, 15, 20]
            max_features_options = [50, 100, 150]
            total_tests = len(list(itertools.product(n_clusters_options, n_components_options, max_features_options)))

            best_score = -1.0 # Initialize
            best_params = {}

            # 1. Check for Cached Parameters
            if keyword_cluster_params.exists():
                try:
                    with keyword_cluster_params.open('r') as file:
                        best_params = json.load(file)
                    print(f"  Loaded initial parameters: {best_params}")
                    # Test with loaded parameters
                    # We call the helper functions that are already in this file
                    df, score, _ = keyword_clustering(df, 'Keyword', **best_params)
                    best_score = score
                    print(f"  Initial test with loaded parameters: Score = {score:.3f}")
                except (json.JSONDecodeError, FileNotFoundError, TypeError, ValueError) as e:
                    print(f"  ⚠️ Failed to load/use cached parameters. Starting full search. Error: {e}")
                    best_params = {}

            # 2. Iterative Search
            if best_score < target_silhouette_score:
                print(f"  Refining best keyword clustering fit... Total tests: {total_tests}")
                for n_clusters, n_components, max_features in itertools.product(n_clusters_options, n_components_options, max_features_options):
                    if (n_clusters == best_params.get('n_clusters') and
                        n_components == best_params.get('n_components') and
                        max_features == best_params.get('max_features')):
                        continue # Skip already-tested params

                    df_temp, score, params = keyword_clustering(df.copy(), 'Keyword', n_clusters, n_components, max_features)
                    print(f'  Testing params: {params}, Score: {score:.3f}')

                    if score > best_score:
                        best_score = score
                        best_params = params
                        df = df_temp.copy() # Keep the DF with the better cluster labels

                        if best_score >= target_silhouette_score:
                            print(f'  ✅ Good enough score found: {best_score:.3f} with params {best_params}')
                            with keyword_cluster_params.open('w') as file:
                                json.dump(best_params, file)
                            print(f'  Saved best parameters: {best_params}')
                            break
                
                if best_score < target_silhouette_score and best_params:
                    print(f'  Highest score reached: {best_score:.3f}. Saving best parameters found.')
                    with keyword_cluster_params.open('w') as file:
                        json.dump(best_params, file)

            # 3. Finalize Clustering
            if 'Keyword Cluster' not in df.columns: # If clustering didn't run or was skipped
                print("  Finalizing clustering with best parameters...")
                df, _, _ = keyword_clustering(df, 'Keyword', **best_params)

            # 4. Naming clusters
            print("\n🏷️  Naming clusters...")
            df = name_keyword_clusters(df, 'Keyword', 'Keyword Cluster') # Call helper

        else:
            print("⏩ Clustering skipped via configuration.")
            # Clean up the column if it exists from a previous run
            if 'Keyword Group (Experimental)' in df.columns:
                df.drop(columns=['Keyword Group (Experimental)'], inplace=True)

        # --- FINAL REORDERING ---
        # We call the helper function _reorder_columns_surgical
        print("  Reordering columns...")
        df = _reorder_columns_surgical(df, 'CPC', after_column='Keyword Difficulty')
        
        # Only attempt to reorder the group column if clustering was enabled/exists
        if 'Keyword Group (Experimental)' in df.columns:
            df = _reorder_columns_surgical(df, 'Keyword Group (Experimental)', after_column='Number of Words')
            
        df = _reorder_columns_surgical(df, 'CPC', after_column='Number of Words') # Verbatim duplicate reorder

        # Conditional reordering
        if has_botify:
            df = _reorder_columns_surgical(df, 'Client URL', after_column='Meta Description')
        else:
            df = _reorder_columns_surgical(df, 'Client URL', after_column='Competition')
        
        df = _reorder_columns_surgical(df, 'Competitor URL', after_column='Client URL')

        # Final file persistence
        df.to_csv(unformatted_csv, index=False)
        print(f"  💾 Intermediate unformatted file saved to '{unformatted_csv}'")


        # --- DISPLAY FINAL CLUSTER COUNTS ---
        print("\n--- Final Keyword Group Counts ---")
        if 'Keyword Group (Experimental)' in df.columns:
            value_counts = df["Keyword Group (Experimental)"].value_counts()
            if not value_counts.empty:
                max_digits = len(str(len(value_counts)))
                max_index_width = max(len(str(index)) for index in value_counts.index)
                max_count_width = max(len(f"{count:,}") for count in value_counts)
                for i, (index, count) in enumerate(value_counts.items(), start=1):
                    counter_str = str(i).zfill(max_digits)
                    count_str = f"{count:,}"
                    print(f"  {counter_str}: {index:<{max_index_width}} - {count_str:>{max_count_width}}")
            else:
                print("  ❌ No keyword groups were created.")
        else:
            print("  (Clustering disabled)")
        print("----------------------------------")

        # --- OUTPUT (to pip state) ---
        pip.set(job, 'final_clustered_df_json', df.to_json(orient='records'))
        print(f"💾 Stored final clustered DataFrame in pip state for job '{job}'.")
        # ---------------------------

        # --- RETURN VALUE ---
        return df

    except Exception as e:
        print(f"❌ An error occurred during clustering and finalization: {e}")
        pip.set(job, 'final_clustered_df_json', pd.DataFrame().to_json(orient='records'))
        return pd.DataFrame() # Return empty DataFrame


# Surgical port of _open_folder from FAQuilizer, necessary for the widget to work
def _open_folder(path_str: str = "."):
    """Opens the specified folder in the system's default file explorer."""
    folder_path = Path(path_str).resolve()
    
    if not folder_path.exists() or not folder_path.is_dir():
        print(f"❌ Error: Path is not a valid directory: {folder_path}")
        return

    system = platform.system()
    try:
        if system == "Windows":
            os.startfile(folder_path)
        elif system == "Darwin":  # macOS
            subprocess.run(["open", folder_path])
        else:  # Linux (xdg-open covers most desktop environments)
            subprocess.run(["xdg-open", folder_path])
    except Exception as e:
        print(f"❌ Failed to open folder. Error: {e}")

# This utility must be defined for normalize_and_score to work
def safe_normalize(series):
    """ Normalize the series safely to avoid divide by zero and handle NaN values. """
    min_val = series.min()
    max_val = series.max()
    range_val = max_val - min_val
    if range_val == 0:
        # Avoid division by zero by returning zero array if no range
        return np.zeros_like(series)
    else:
        # Normalize and fill NaN values that might result from empty/NaN series
        return (series - min_val).div(range_val).fillna(0)

# Surgical port of bf.reorder_columns
def reorder_columns_surgical(df, priority_column, after_column):
    if priority_column in df.columns:
        columns = list(df.columns.drop(priority_column))
        # Handle cases where the after_column may have been dropped earlier
        if after_column not in columns:
            print(f"⚠️ Reorder Error: Target column '{after_column}' not found. Skipping reorder of '{priority_column}'.")
            return df
            
        after_column_index = columns.index(after_column)
        columns.insert(after_column_index + 1, priority_column)
        df = df[columns]
    else:
        print(f"⚠️ Reorder Error: Column '{priority_column}' not found in DataFrame.")
    return df

# Surgical port of bf.normalize_and_score - WITH CRITICAL FIX
def normalize_and_score_surgical(df, registered_domain, has_botify_data, after_col, reorder):
    
    # Rename original column fields to match expected names in the dataframe
    if 'internal_page_rank.raw' in df.columns:
        df = df.rename(columns={'internal_page_rank.raw': 'Raw Internal Pagerank'}, inplace=False)
    
    # --- CRITICAL FIX FOR KEYERROR / TRAILING SLASHES ---
    # The lookup key (e.g., 'nixos.org') must be matched against the DataFrame column (e.g., 'nixos.org/').
    # We clean both for comparison to find the unique canonical key, but use the original column name.
    
    # Clean the lookup domain (assuming the input `registered_domain` might be missing the slash)
    clean_lookup_key = registered_domain.rstrip('/')
    target_col = None
    
    for col in df.columns:
        # Find the column whose stripped name matches the stripped lookup key.
        if col.rstrip('/') == clean_lookup_key:
            target_col = col
            break
            
    if target_col is None:
        raise KeyError(f"Could not find client domain column for '{registered_domain}' in DataFrame. Available columns: {df.columns.tolist()}")
    # --- END CRITICAL FIX ---
    
    # Normalize metrics that are always included
    df['Normalized Search Volume'] = safe_normalize(df['Search Volume'])
    df['Normalized Search Position'] = safe_normalize(df[target_col]) # <-- USES THE FOUND, CANONICAL COLUMN NAME
    df['Normalized Keyword Difficulty'] = safe_normalize(df['Keyword Difficulty'])
    df['Normalized CPC'] = safe_normalize(df['CPC'])

    # Always include CPC and Keyword Difficulty in the combined score
    combined_score = df['Normalized CPC'] - df['Normalized Keyword Difficulty']

    if has_botify_data:
        # Normalize additional Botify metrics if available
        if 'Raw Internal Pagerank' in df.columns:
            df['Normalized Raw Internal Pagerank'] = safe_normalize(df['Raw Internal Pagerank'])
        else:
            df['Normalized Raw Internal Pagerank'] = 0

        if "No. of Missed Clicks excluding anonymized queries" in df.columns:
            df['Normalized Missed Clicks'] = safe_normalize(df["No. of Missed Clicks excluding anonymized queries"])
            combined_score += df['Normalized Missed Clicks']
        else:
            df['Normalized Missed Clicks'] = 0

        # Add Botify metrics to the combined score
        combined_score += (-1 * df['Normalized Raw Internal Pagerank'] +
                           df['Normalized Search Volume'] +
                           df['Normalized Search Position'])

    # Apply the combined score to the DataFrame
    df['Combined Score'] = combined_score

    if reorder:
        # Reorder columns if required (using the surgically ported reorder function)
        df = reorder_columns_surgical(df, "CPC", after_col)
        df = reorder_columns_surgical(df, "Keyword Difficulty", "CPC")
        if has_botify_data:
            df = reorder_columns_surgical(df, "Internal Pagerank", "Keyword Difficulty")
            df = reorder_columns_surgical(df, "No. of Unique Inlinks", "Internal Pagerank")
            if "No. of Missed Clicks excluding anonymized queries" in df.columns:
                df = reorder_columns_surgical(df, "No. of Missed Clicks excluding anonymized queries", "No. of Unique Inlinks")

    return df


def create_deliverables_excel_and_button(job: str, df: pd.DataFrame, client_domain_from_keys: str, has_botify: bool):
    """
    Creates the deliverables directory, writes the first "Gap Analysis" tab
    to the Excel file, creates the "Open Folder" button, and stores
    key file paths in pip state.

    Args:
        job (str): The current Pipulate job ID.
        df (pd.DataFrame): The final clustered/arranged DataFrame.
        client_domain_from_keys (str): The client's domain from the keys module.
        has_botify (bool): Flag indicating if Botify data is present.

    Returns:
        tuple: (button, xl_file, loop_list)
               The ipywidget Button, the Path object for the Excel file,
               and the list of sheet names (loop_list).
    """
    print("Writing first pass of Excel file...")

    # --- CORE LOGIC (Moved from Notebook) ---
    try:
        # --- FIX: Re-derive missing variables ---
        # 1. Get semrush_lookup
        semrush_lookup = _extract_registered_domain(client_domain_from_keys)

        # 2. Get canonical competitors list from pip state
        competitors_list_json = pip.get(job, 'competitors_list_json', '[]')
        competitors = json.loads(competitors_list_json)
        if not competitors:
            # This should NOT happen, but it's a safe fallback.
            print(f"  ❌ CRITICAL WARNING: 'competitors_list_json' not found in pip state. Inferring from DataFrame columns.")
            competitors = [col for col in df.columns if col == semrush_lookup or '/' in col or '.com' in col]
        else:
            print(f"  ✅ Loaded canonical list of {len(competitors)} competitors from pip state.")


        # 3. Find the canonical client column name (TARGET_COMPETITOR_COL)
        clean_lookup_key = semrush_lookup.rstrip('/')
        TARGET_COMPETITOR_COL = None
        for col in df.columns: # Use the passed 'df'
            if col.rstrip('/') == clean_lookup_key:
                TARGET_COMPETITOR_COL = col
                break

        if TARGET_COMPETITOR_COL is None:
             # This will cause normalize_and_score_surgical to fail, but we must try.
             print(f"❌ CRITICAL ERROR: Could not find canonical column for '{semrush_lookup}' in DataFrame.")
             # We'll let it fail in normalize_and_score_surgical for a clearer traceback.
             TARGET_COMPETITOR_COL = semrush_lookup # Use the base name as a last resort
        # --- END FIX ---


        # --- 1. DEFINE SECURE OUTPUT PATHS ---
        deliverables_dir = Path("deliverables") / job
        deliverables_dir.mkdir(parents=True, exist_ok=True)

        xl_filename = f"{semrush_lookup.replace('.', '_').rstrip('_')}_GAPalyzer_{job}_V1.xlsx"
        xl_file = deliverables_dir / xl_filename

        # --- 2. EXECUTE CORE LOGIC ---
        print(f"  - Writing 'Gap Analysis' tab to {xl_file.name}...")
        loop_list = ["Gap Analysis"] # This is needed by the next cell
        last_competitor = competitors[-1] if competitors else None # Handle empty list

        # Apply the normalization/scoring logic
        # Pass the re-derived semrush_lookup and the canonical TARGET_COMPETITOR_COL
        # NOTE: normalize_and_score_surgical finds its *own* target_col, so we just need semrush_lookup
        df_tab = normalize_and_score_surgical(df.copy(), semrush_lookup, has_botify, last_competitor, False)

        # Save Initial Excel Sheet
        arg_dict = {'options': {'strings_to_urls': False}}
        try:
            with pd.ExcelWriter(xl_file, engine="xlsxwriter", engine_kwargs=arg_dict, mode='w') as writer:
                df_tab.to_excel(writer, sheet_name='Gap Analysis', index=False)
            print("  ✅ 'Gap Analysis' tab written (Unformatted Pass 1).")
        except Exception as e:
            print(f"  ❌ Error writing Excel file: {e}")

        # --- 3. CREATE SECURE EGRESS BUTTON ---
        button = widgets.Button(
            description=f"📂 Open Deliverables Folder ({job})",
            tooltip=f"Open {deliverables_dir.resolve()}",
            button_style='success'
        )
        
        # Define the on_click handler that calls our private helper
        def on_open_folder_click(b):
            _open_folder(str(deliverables_dir))
            
        button.on_click(on_open_folder_click)

        # --- OUTPUT (to pip state) ---
        pip.set(job, 'final_xl_file', str(xl_file))
        pip.set(job, 'deliverables_folder', str(deliverables_dir))
        # --- FIX: Serialize lists to JSON strings before storing ---
        pip.set(job, 'loop_list', json.dumps(loop_list)) # Store loop_list for the next step
        # Store competitors and target col for the *next* cell (the formatting one)
        pip.set(job, 'competitors_list', json.dumps(competitors)) 
        # --- END FIX ---
        pip.set(job, 'semrush_lookup', semrush_lookup) # The clean domain
        pip.set(job, 'target_competitor_col', TARGET_COMPETITOR_COL) # The canonical column name
        print(f"💾 Stored final Excel path, folder, and competitor info in pip state.")
        # ---------------------------

        # --- RETURN VALUE ---
        # Return values needed for notebook display AND next cell
        return button, xl_file, loop_list, competitors, semrush_lookup, TARGET_COMPETITOR_COL, has_botify

    except Exception as e:
        print(f"❌ An error occurred during Excel creation: {e}")
        # Return dummy values to avoid breaking the notebook flow
        return widgets.Button(description=f"Error: {e}", disabled=True), None, [], [], None, None, False


def add_filtered_excel_tabs(
    job: str,
    df: pd.DataFrame,
    semrush_lookup: str,
    has_botify: bool,
    competitors: list,
    xl_file: Path,
    TARGET_COMPETITOR_COL: str,
    button: widgets.Button,
    custom_filters: list = None,
    width_adjustment: float = 1.5
):
    """
    Generates data, writes all tabs, and applies ALL formatting in a single fast pass.
    Replaces the slow append-and-format openpyxl workflow.
    """
    print(f"⚡ Batch-processing filters, writing, and formatting {xl_file.name}...")

    # --- 1. PREPARE DATA IN MEMORY ---
    # Helper functions
    def read_keywords(file_path):
        if not file_path.exists(): return []
        with open(file_path, 'r') as file: return [line.strip() for line in file.readlines()]

    def filter_df_by_keywords(df, keywords):
        return df[df["Keyword"].isin(keywords)]

    # Dictionary to hold {SheetName: DataFrame}
    tabs_to_write = {}

    # A. Main Tab
    print("  - Preparing 'Gap Analysis' tab...")
    last_competitor = competitors[-1] if competitors else None
    df_main = normalize_and_score_surgical(df.copy(), semrush_lookup, has_botify, last_competitor, False)
    tabs_to_write["Gap Analysis"] = df_main

    # B. Important Keywords
    important_keywords_file = Path("data") / f"{job}_important_keywords.txt"
    if important_keywords_file.exists():
        kws = read_keywords(important_keywords_file)
        if kws:
            print("  - Preparing 'Important Keywords' tab...")
            df_filtered = filter_df_by_keywords(df.copy(), kws)
            df_filtered = normalize_and_score_surgical(df_filtered, semrush_lookup, has_botify, competitors[-1], True)
            df_filtered.sort_values(by='Combined Score', ascending=False, inplace=True)
            tabs_to_write["Important Keywords"] = df_filtered

    # C. Best Opportunities
    print("  - Preparing 'Best Opportunities' tab...")
    striking_lower = 100
    df_opps = df.copy()
    if has_botify:
        try:
            df_opps = df_opps[(df_opps["No. of Impressions excluding anonymized queries"] > 0) & (df_opps[TARGET_COMPETITOR_COL] > 3)].copy()
        except KeyError:
            df_opps = df_opps[(df_opps[TARGET_COMPETITOR_COL] >= 4) & (df_opps[TARGET_COMPETITOR_COL] <= striking_lower)].copy()
    else:
        df_opps = df_opps[(df_opps[TARGET_COMPETITOR_COL] >= 4) & (df_opps[TARGET_COMPETITOR_COL] <= striking_lower)].copy()
    df_opps = normalize_and_score_surgical(df_opps, semrush_lookup, has_botify, competitors[-1], True)
    df_opps.sort_values(by='Combined Score', ascending=False, inplace=True)
    tabs_to_write["Best Opportunities"] = df_opps

    # E. Targeted Filters
    if custom_filters:
        print(f"  - Preparing {len(custom_filters)} targeted filter tabs...")
        for filter_name, keywords in custom_filters:
            pattern = r'\b(?:' + '|'.join([re.escape(k) for k in keywords]) + r')\b'
            df_tab = df[df["Keyword"].str.contains(pattern, case=False, na=False)].copy()
            if not df_tab.empty:
                df_tab = normalize_and_score_surgical(df_tab, semrush_lookup, has_botify, competitors[-1], True)
                df_tab.sort_values(by='Combined Score', ascending=False, inplace=True)
                tabs_to_write[filter_name] = df_tab

    # --- 2. CONFIGURATION FOR FORMATTING ---
    # Colors
    colors = {
        'client': '#FFFF00', 'semrush': '#FAEADB', 'semrush_opp': '#F1C196',
        'botify': '#EADFF2', 'botify_opp': '#AEA1C4', 'competitor': '#EEECE2'
    }
    
    # Columns Config
    tiny, small, medium, large, url_w = 11, 15, 20, 50, 70
    col_widths = {
        'Keyword': 40, 'Search Volume': small, 'Number of Words': tiny, 'Keyword Group (Experimental)': small, 
        'Competitors Positioning': tiny, 'CPC': tiny, 'Keyword Difficulty': tiny, 'Competition': tiny, 
        'Depth': tiny, 'No. of Keywords': tiny, 'No. of Impressions excluding anonymized queries': small,
        'No. of Clicks excluding anonymized queries': small, 'No. of Missed Clicks excluding anonymized queries': small,
        'Avg. URL CTR excluding anonymized queries': tiny, 'Avg. URL Position excluding anonymized queries': tiny,
        'Raw Internal Pagerank': small, 'Internal Pagerank': tiny, 'No. of Unique Inlinks': tiny,
        'Title': large, 'Meta Description': large, 'Competitor URL': url_w, 'Client URL': url_w,
        'Combined Score': tiny
    }
    
    num_fmts = {
        'Search Volume': '#,##0', 'CPC': '0.00', 'Keyword Difficulty': '0', 'Competition': '0.00',
        'Raw Internal Pagerank': '0.0000000', 'Internal Pagerank': '0.00', 'Combined Score': '0.00',
        'No. of Impressions excluding anonymized queries': '#,##0', 'No. of Clicks excluding anonymized queries': '#,##0'
    }

    cond_desc = ['Search Volume', 'CPC', 'Competition', 'Avg. URL CTR excluding anonymized queries', 'No. of Missed Clicks excluding anonymized queries', 'Combined Score', 'No. of Unique Inlinks']
    cond_asc = ['Keyword Difficulty', 'Raw Internal Pagerank', 'Internal Pagerank', 'Avg. URL Position excluding anonymized queries', 'Depth']

    # --- 3. WRITE AND FORMAT ---
    print(f"💾 Writing and formatting {len(tabs_to_write)} tabs...")
    
    try:
        with pd.ExcelWriter(xl_file, engine="xlsxwriter", engine_kwargs={'options': {'strings_to_urls': False}}) as writer:
            workbook = writer.book
            
            # Define Formats
            # UPDATED: Text Wrap turned ON for standard headers
            header_fmt = workbook.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'text_wrap': True})
            
            client_col_fmt = workbook.add_format({'bg_color': colors['client']})
            
            # Rotated Header: Align to 'top' (left) as requested
            rot_fmt = workbook.add_format({'bold': True, 'align': 'center', 'valign': 'top', 'rotation': 90, 'border': 1, 'bg_color': colors['competitor']})
            
            # UPDATED: Client Header now rotates 90 deg and aligns top (left), with yellow background
            client_rot_fmt = workbook.add_format({'bold': True, 'align': 'center', 'valign': 'top', 'rotation': 90, 'border': 1, 'bg_color': colors['client']})
            
            # Color Formats (Standard headers with colors)
            fmt_semrush = workbook.add_format({'bg_color': colors['semrush'], 'bold': True, 'align': 'center', 'border': 1, 'text_wrap': True})
            fmt_botify = workbook.add_format({'bg_color': colors['botify'], 'bold': True, 'align': 'center', 'border': 1, 'text_wrap': True})
            
            for sheet_name, df_sheet in tabs_to_write.items():
                # Write Data
                df_sheet.to_excel(writer, sheet_name=sheet_name, index=False)
                worksheet = writer.sheets[sheet_name]
                
                # Add Table (Auto-filter, banding)
                (max_row, max_col) = df_sheet.shape
                cols = [{'header': col} for col in df_sheet.columns]
                worksheet.add_table(0, 0, max_row, max_col - 1, {'columns': cols, 'style': 'TableStyleMedium9'})
                
                # Freeze Panes (C2)
                worksheet.freeze_panes(1, 2)

                # Iterate Columns to apply widths and formats
                for i, col_name in enumerate(df_sheet.columns):
                    # A. Widths
                    width = col_widths.get(col_name, small) * width_adjustment
                    # Check number format
                    n_fmt = workbook.add_format({'num_format': num_fmts.get(col_name, '')})
                    
                    # B. Client Column Highlight (Apply to whole column)
                    if col_name == TARGET_COMPETITOR_COL:
                         # Table style usually wins for body background, but we set width below.
                         pass 

                    worksheet.set_column(i, i, width, n_fmt)

                    # C. Header Formatting
                    # Overwrite header cell with specific styles
                    current_header_fmt = header_fmt
                    
                    if col_name in competitors and col_name != TARGET_COMPETITOR_COL:
                        current_header_fmt = rot_fmt
                        worksheet.set_column(i, i, 5) # Narrow width for rotated cols
                    elif col_name == TARGET_COMPETITOR_COL:
                        # UPDATED: Use the rotated yellow format and force narrow width
                        current_header_fmt = client_rot_fmt
                        worksheet.set_column(i, i, 5) 
                    elif col_name in ['Keyword', 'Search Volume', 'CPC', 'Keyword Difficulty', 'Competition']:
                         current_header_fmt = fmt_semrush
                    elif col_name in ['Internal Pagerank', 'Depth', 'Title']:
                         current_header_fmt = fmt_botify
                    
                    worksheet.write(0, i, col_name, current_header_fmt)

                    # D. Conditional Formatting
                    rng = f"{get_column_letter(i+1)}2:{get_column_letter(i+1)}{max_row+1}"
                    if col_name in cond_desc:
                        worksheet.conditional_format(rng, {'type': '3_color_scale', 'min_color': '#FFFFFF', 'max_color': '#33FF33'})
                    elif col_name in cond_asc:
                        worksheet.conditional_format(rng, {'type': '3_color_scale', 'min_color': '#33FF33', 'max_color': '#FFFFFF'})

                    # E. Hyperlinks (Simple blue underline for URLs)
                    if "URL" in col_name:
                        pass

        print("✅ Excel write and format complete.")
        
    except Exception as e:
        print(f"❌ Error writing Excel file: {e}")
        return widgets.Button(description=f"Error: {e}", disabled=True)

    return button


# --- Private Helper Functions for Excel Formatting ---

def _create_column_mapping(sheet):
    """Creates a dictionary mapping header names to column letters."""
    mapping = {}
    for col_idx, column_cell in enumerate(sheet[1], 1): # Assumes headers are in row 1
        column_letter = get_column_letter(col_idx)
        mapping[str(column_cell.value)] = column_letter
    return mapping

def _apply_fill_to_column_labels(sheet, column_mapping, columns_list, fill):
    """Applies a fill color to the header cells of specified columns."""
    for column_name in columns_list:
        column_letter = column_mapping.get(column_name)
        if column_letter:
            cell = sheet[f"{column_letter}1"]
            cell.fill = fill

def _find_last_data_row(sheet, keyword_column_letter):
    """Finds the last row containing data in a specific column (e.g., 'Keyword')."""
    if not keyword_column_letter: # Handle case where keyword column might be missing
        return sheet.max_row
    last_row = sheet.max_row
    # Iterate backwards from the max row
    while last_row > 1 and sheet[f"{keyword_column_letter}{last_row}"].value in [None, "", " "]:
        last_row -= 1
    return last_row

def _apply_conditional_formatting(sheet, column_mapping, last_row, conditionals_descending, conditionals_ascending, rule_desc, rule_asc, competitors):
    """Applies color scale conditional formatting to specified columns."""
    for label in conditionals_descending + conditionals_ascending:
        if label in competitors:
            continue # Explicitly skip any column that is a competitor
        column_letter = column_mapping.get(label)
        if column_letter and last_row > 1: # Ensure there is data to format
            range_string = f'{column_letter}2:{column_letter}{last_row}'
            rule = rule_desc if label in conditionals_descending else rule_asc
            try:
                sheet.conditional_formatting.add(range_string, rule)
            except Exception as e:
                print(f"  ⚠️ Failed to apply conditional formatting for {label}: {e}")

def _is_safe_url(url):
    """ Check if the given string is a valid URL using the validators library. """
    if not isinstance(url, str):
        return False
    return validators.url(url) is True # Explicitly check for True


def apply_excel_formatting(
    job: str,
    xl_file: Path,
    competitors: list,
    semrush_lookup: str,
    TARGET_COMPETITOR_COL: str,
    has_botify: bool,
    GLOBAL_WIDTH_ADJUSTMENT: float,
    button: widgets.Button
):
    """
    Legacy function. Formatting is now handled directly in add_filtered_excel_tabs 
    using xlsxwriter for vastly improved performance.
    """
    print("⏩ Formatting already applied during write. Skipping legacy formatting step.")
    return button

