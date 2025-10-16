# secretsauce.py (version 3.0 - Refactored Workflow)
# This module contains the implementation details for a 1-to-many AI enrichment workflow.

from pipulate import pip
import google.generativeai as genai
import requests
from bs4 import BeautifulSoup
import pandas as pd
from io import StringIO
import json
from sqlitedict import SqliteDict
import asyncio
import nbformat
from pathlib import Path
import re

from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo


# --- CONFIGURATION ---
CACHE_DB_FILE = "url_cache.sqlite"
EXTRACTED_DATA_CSV = "_step_extract_output.csv"
AI_LOG_CSV = "_step_ai_log_output.csv" # NEW: Filename for the AI output log

# Pipulate step names
API_KEY_STEP = "api_key"
URL_LIST_STEP = "url_list"
EXTRACTED_DATA_STEP = "extracted_data"
FAQ_DATA_STEP = "faq_data"
FINAL_DATAFRAME_STEP = "final_dataframe"
EXPORT_FILE_STEP = "export_file_path"


def _get_prompt_from_notebook(notebook_filename="FAQuilizer.ipynb"):
    """Parses a notebook file to extract the prompt from the 'prompt-input' tagged cell."""
    try:
        notebook_path = Path(__file__).parent / notebook_filename
        with open(notebook_path, 'r', encoding='utf-8') as f:
            nb = nbformat.read(f, as_version=4)
        
        for cell in nb.cells:
            if "prompt-input" in cell.metadata.get("tags", []):
                return cell.source
        return None # Return None if the tag isn't found
    except Exception as e:
        print(f"⚠️ Could not read prompt from notebook: {e}")
        return None


def _get_urls_from_notebook(notebook_filename="FAQuilizer.ipynb"):
    """Parses a notebook file to extract URLs from the 'url-list-input' tagged cell."""
    try:
        # Assuming the notebook is in the same directory as this script
        notebook_path = Path(__file__).parent / notebook_filename
        with open(notebook_path, 'r', encoding='utf-8') as f:
            nb = nbformat.read(f, as_version=4)
        
        for cell in nb.cells:
            if "url-list-input" in cell.metadata.get("tags", []):
                urls_raw = cell.source
                urls = [
                    line.split('#')[0].strip() 
                    for line in urls_raw.splitlines() 
                    if line.strip() and not line.strip().startswith('#')
                ]
                return urls
        return []
    except Exception as e:
        print(f"⚠️ Could not read URLs from notebook: {e}")
        return []


async def scrape_and_extract(job: str, 
                             headless: bool = True, 
                             verbose: bool = False, 
                             stealth: bool = True, 
                             persistent: bool = False, 
                             profile_name: str = "default", 
                             delay_range: tuple = (5, 10)):
    """
    Scrapes each URL using pip.scrape(), leveraging cached data if available,
    and immediately parses the HTML to extract key SEO data.
    """
    print("🚀 Starting browser-based scraping and extraction...")
    
    # --- Read fresh URLs from the notebook and update the state ---
    fresh_urls = _get_urls_from_notebook()
    if fresh_urls:
        print(f"✨ Found {len(fresh_urls)} URLs in the notebook.")
        pip.set(job, URL_LIST_STEP, fresh_urls)
    # --------------------------------------------------------------------

    urls_to_process = pip.get(job, URL_LIST_STEP, [])
    if not urls_to_process:
        print("❌ No URLs to process. Please add them to the 'url-list-input' cell in your notebook.")
        return

    extracted_data = []

    for i, url in enumerate(urls_to_process):
        # The logging is now cleaner, showing a distinct message for cached items.
        # The core processing logic remains the same.
        
        # Apply delay only AFTER the first request to avoid an unnecessary initial wait
        current_delay_range = delay_range if i > 0 else None

        try:
            scrape_result = await pip.scrape(
                url=url,
                take_screenshot=True,
                headless=headless,
                verbose=verbose,
                stealth=stealth,
                persistent=persistent,
                profile_name=profile_name,
                delay_range=current_delay_range
            )
            
            # --- AESTHETIC LOGGING UPDATE ---
            is_cached = scrape_result.get("cached", False)
            if is_cached:
                print(f"  -> ✅ Cached [{i+1}/{len(urls_to_process)}] Using data for: {url}")
            else:
                print(f"  -> 👁️  Scraped [{i+1}/{len(urls_to_process)}] New data for: {url}")


            if not scrape_result.get("success"):
                if verbose:
                    print(f"  -> ❌ Scrape failed: {scrape_result.get('error')}")
                continue

            dom_path = scrape_result.get("looking_at_files", {}).get("rendered_dom")
            if not dom_path:
                if verbose:
                    print(f"  -> ⚠️ Scrape succeeded, but no DOM file was found.")
                continue

            with open(dom_path, 'r', encoding='utf-8') as f:
                html_content = f.read()
            
            soup = BeautifulSoup(html_content, 'html.parser')
            title = soup.title.string.strip() if soup.title else "No Title Found"
            meta_desc_tag = soup.find('meta', attrs={'name': 'description'})
            meta_description = meta_desc_tag['content'].strip() if meta_desc_tag else ""
            h1s = [h1.get_text(strip=True) for h1 in soup.find_all('h1')]
            h2s = [h2.get_text(strip=True) for h2 in soup.find_all('h2')]
            
            extracted_data.append({
                'url': url, 'title': title, 'meta_description': meta_description,
                'h1s': h1s, 'h2s': h2s
            })
            # No need for a verbose check here, the new logging is always informative.

        except Exception as e:
            print(f"  -> ❌ A critical error occurred while processing {url}: {e}")

    pip.set(job, EXTRACTED_DATA_STEP, extracted_data)
    print(f"✅ Scraping and extraction complete for {len(extracted_data)} URLs.")


def extract_webpage_data(job: str):
    """Reads from cache, extracts key SEO elements, and saves to CSV."""
    urls_to_process = pip.get(job, URL_LIST_STEP, [])
    extracted_data = []
    print(f"🔍 Extracting SEO elements for {len(urls_to_process)} URLs...")
    with SqliteDict(CACHE_DB_FILE) as cache:
        for url in urls_to_process:
            response = cache.get(url)
            if not response or not isinstance(response, requests.Response):
                print(f"  -> ⏭️ Skipping {url} (no valid cache entry).")
                continue
            print(f"  -> Parsing {url}...")
            soup = BeautifulSoup(response.content, 'html.parser')
            title = soup.title.string.strip() if soup.title else "No Title Found"
            meta_desc_tag = soup.find('meta', attrs={'name': 'description'})
            meta_description = meta_desc_tag['content'].strip() if meta_desc_tag else ""
            h1s = [h1.get_text(strip=True) for h1 in soup.find_all('h1')]
            h2s = [h2.get_text(strip=True) for h2 in soup.find_all('h2')]
            extracted_data.append({
                'url': url, 'title': title, 'meta_description': meta_description,
                'h1s': h1s, 'h2s': h2s
            })
    pip.set(job, EXTRACTED_DATA_STEP, extracted_data)
    try:
        df = pd.DataFrame(extracted_data)
        df.to_csv(EXTRACTED_DATA_CSV, index=False)
        print(f"✅ Extraction complete. Intermediate data saved to '{EXTRACTED_DATA_CSV}'")
    except Exception as e:
        print(f"⚠️ Could not save intermediate CSV: {e}")


# -----------------------------------------------------------------------------
# NEW REFACTORED WORKFLOW: Stack 'Em, FAQ 'Em, Rack 'Em
# -----------------------------------------------------------------------------

def stack_em(job: str) -> pd.DataFrame:
    """
    Loads pre-scraped and extracted data for a job into a DataFrame.
    This is the "Stack 'Em" step.
    """
    print("📊 Stacking pre-extracted data into a DataFrame...")
    extracted_data = pip.get(job, EXTRACTED_DATA_STEP, [])
    if not extracted_data:
        print("❌ No extracted data found. Please run `scrape_and_extract` first.")
        return pd.DataFrame()

    df = pd.DataFrame(extracted_data)
    print(f"✅ Stacked {len(df)} pages into the initial DataFrame.")
    return df

def ai_faq_em(job: str, debug: bool = False) -> pd.DataFrame:
    """
    Enriches scraped data with AI-generated FAQs, using a JSON file for robust caching
    to avoid re-processing URLs. This is the "FAQ 'Em" step.
    """
    import os
    import json
    from pathlib import Path
    import google.generativeai as genai
    import re

    # --- 1. Define Cache Path ---
    # The script runs from the Notebooks directory, so the path is relative to that.
    cache_dir = Path("data")
    cache_dir.mkdir(parents=True, exist_ok=True)
    cache_file = cache_dir / f"faq_cache_{job}.json"

    # --- 2. Load Data ---
    extracted_data = pip.get(job, EXTRACTED_DATA_STEP, [])
    if not extracted_data:
        print("❌ No extracted data found. Please run `scrape_and_extract` first.")
        return pd.DataFrame()

    faq_data = []
    if cache_file.exists():
        try:
            raw_content = cache_file.read_text(encoding='utf-8')
            if raw_content.strip():
                faq_data = json.loads(raw_content)
                print(f"✅ Loaded {len(faq_data)} FAQs from cache.")
        except (json.JSONDecodeError, IOError) as e:
            print(f"⚠️ Could not load cache file. Starting fresh. Error: {e}")
    
    processed_urls = {item.get('url') for item in faq_data}
    print(f"🧠 Generating FAQs for {len(extracted_data)} pages... ({len(processed_urls)} already cached)")

    # --- 3. Get Prompt & Configure AI ---
    user_prompt_instructions = _get_prompt_from_notebook()
    if not user_prompt_instructions:
        print("❌ Error: Prompt not found in 'prompt-input' cell of the notebook.")
        return pd.DataFrame(faq_data)
        
    system_prompt_wrapper = '''
Your task is to analyze webpage data and generate a structured JSON object.
Your output must be **only a single, valid JSON object inside a markdown code block** and nothing else. Adherence to the schema is critical.

--- START USER INSTRUCTIONS ---

{user_instructions}

--- END USER INSTRUCTIONS ---

**Input Data:**

--- WEBPAGE DATA BEGIN ---
{webpage_data}
--- WEBPAGE DATA END ---

**Final Instructions:**

Based *only* on the provided webpage data and the user instructions, generate the requested data.
Remember, your entire output must be a single JSON object in a markdown code block. Do not include any text or explanation outside of this block.

The JSON object must conform to the following schema:

{{
  "faqs": [
    {{
      "priority": "integer (1-5, 1 is highest)",
      "question": "string (The generated question)",
      "target_intent": "string (What is the user's goal in asking this?)",
      "justification": "string (Why is this a valuable question to answer? e.g., sales, seasonal, etc.)"
    }}
  ]
}}
'''
    # The API key is configured via pip.api_key() in the notebook.
    # This function assumes that has been run.
    
    # --- 4. Process Loop ---
    try:
        model = genai.GenerativeModel('models/gemini-2.5-flash')
        for index, webpage_data_dict in enumerate(extracted_data):
            url = webpage_data_dict.get('url')
            if url in processed_urls:
                print(f"  -> ✅ Skip: URL already cached: {url}")
                continue

            print(f"  -> 🤖 AI Call: Processing URL {index+1}/{len(extracted_data)}: {url}")
            
            try:
                webpage_data_str = json.dumps(webpage_data_dict, indent=2)
                full_prompt = system_prompt_wrapper.format(
                    user_instructions=user_prompt_instructions,
                    webpage_data=webpage_data_str
                )
                
                if debug:
                    print("\n--- PROMPT ---")
                    print(full_prompt)
                    print("--- END PROMPT ---\n")

                ai_response = model.generate_content(full_prompt)
                response_text = ai_response.text.strip()
                # New robust JSON cleaning
                clean_json = response_text
                if clean_json.startswith("```json"):
                    clean_json = clean_json[7:]
                if clean_json.startswith("```"):
                    clean_json = clean_json[3:]
                if clean_json.endswith("```"):
                    clean_json = clean_json[:-3]
                clean_json = clean_json.strip()

                faq_json = json.loads(clean_json)
                
                new_faqs_for_url = []
                for faq in faq_json.get('faqs', []):
                    new_faqs_for_url.append({
                        'url': url,
                        'title': webpage_data_dict.get('title'),
                        'priority': faq.get('priority'),
                        'question': faq.get('question'),
                        'target_intent': faq.get('target_intent'),
                        'justification': faq.get('justification')
                    })
                
                if new_faqs_for_url:
                    faq_data.extend(new_faqs_for_url)
                    processed_urls.add(url)
                    print(f"  -> ✅ Success: Generated {len(new_faqs_for_url)} new FAQs for {url}.")

            except json.JSONDecodeError as e:
                print(f"  -> ❌ JSON Decode Error for {url}: {e}")
                print(f"  -> Raw AI Response:\n---\n{response_text}\n---")
                continue # Skip to the next URL
            except Exception as e:
                print(f"  -> ❌ AI call failed for {url}: {e}")
                continue

    except KeyboardInterrupt:
        print("\n🛑 Execution interrupted by user.")
    except Exception as e:
        print(f"❌ An error occurred during FAQ generation: {e}")
    finally:
        print("\n💾 Saving progress to cache...")
        try:
            with open(cache_file, 'w', encoding='utf-8') as f:
                json.dump(faq_data, f, indent=2)
            print(f"✅ Save complete. {len(faq_data)} total FAQs in cache.")
        except Exception as e:
            print(f"❌ Error saving cache in `finally` block: {e}")

    print("✅ FAQ generation complete.")
    pip.set(job, FAQ_DATA_STEP, faq_data)
    return pd.DataFrame(faq_data)

def rack_em(df: pd.DataFrame) -> pd.DataFrame:
    """
    Pivots and reorders the long-format FAQ data into a wide-format DataFrame.
    Each URL gets one row, with columns for each of its generated FAQs.
    This is the "Rack 'Em" step.
    """
    if df.empty:
        print("⚠️ DataFrame is empty, skipping the pivot.")
        return pd.DataFrame()

    print("🔄 Racking the data into its final wide format...")

    # 1. Create a unique identifier for each FAQ within a URL group.
    df['faq_num'] = df.groupby('url').cumcount() + 1

    # 2. Set index and unstack to pivot the data.
    pivoted_df = df.set_index(['url', 'title', 'faq_num']).unstack(level='faq_num')

    # 3. Flatten the multi-level column index.
    pivoted_df.columns = [f'{col[0]}_{col[1]}' for col in pivoted_df.columns]
    pivoted_df = pivoted_df.reset_index()

    # --- NEW: Reorder columns for readability ---
    print("🤓 Reordering columns for logical grouping...")

    # Identify the static columns
    static_cols = ['url', 'title']
    
    # Dynamically find the stems (e.g., 'priority', 'question')
    # This makes the code adaptable to different column names
    stems = sorted(list(set(
        col.rsplit('_', 1)[0] for col in pivoted_df.columns if '_' in col
    )))

    # Dynamically find the max FAQ number
    num_faqs = max(
        int(col.rsplit('_', 1)[1]) for col in pivoted_df.columns if col.rsplit('_', 1)[-1].isdigit()
    )

    # Build the new column order
    new_column_order = static_cols.copy()
    for i in range(1, num_faqs + 1):
        for stem in stems:
            new_column_order.append(f'{stem}_{i}')
    
    # Apply the new order
    reordered_df = pivoted_df[new_column_order]
    
    print("✅ Data racked and reordered successfully.")
    return reordered_df


def display_results_log(job: str):
    """
    MODIFIED: Displays the FAQ log AND saves it to an intermediate CSV.
    """
    print("📊 Displaying raw FAQ log...")
    faq_data = pip.get(job, FAQ_DATA_STEP, [])
    if not faq_data:
        print("No FAQ data to display. Please run the previous steps.")
        return
    
    df = pd.DataFrame(faq_data)
    
    # NEW: Save the second "factory floor" file for transparency.
    try:
        df.to_csv(AI_LOG_CSV, index=False)
        print(f"  -> Intermediate AI log saved to '{AI_LOG_CSV}'")
    except Exception as e:
        print(f"⚠️ Could not save AI log CSV: {e}")

    pip.set(job, FINAL_DATAFRAME_STEP, df.to_json(orient='records'))
    with pd.option_context('display.max_rows', None, 'display.max_colwidth', 80):
        display(df)

def export_to_excel(job: str):
    """Exports the final DataFrame to a formatted Excel file."""
    print("📄 Exporting data to Excel...")
    final_json = pip.get(job, FINAL_DATAFRAME_STEP)
    if not final_json:
        print("❌ No final data found to export. Please run the 'display_results' step first.")
        return
    df_final = pd.read_json(StringIO(final_json))
    output_filename = f"{job}_output.xlsx"
    try:
        with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
            df_final.to_excel(writer, index=False, sheet_name='Faquillizer_Data')
            worksheet = writer.sheets['Faquillizer_Data']
            for column in worksheet.columns:
                max_length = max((df_final[column[0].value].astype(str).map(len).max(), len(str(column[0].value))))
                adjusted_width = (max_length + 2) if max_length < 80 else 80
                worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
        pip.set(job, EXPORT_FILE_STEP, output_filename)
        print(f"✅ Success! Data exported to '{output_filename}'")
    except Exception as e:
        print(f"❌ Failed to export to Excel: {e}")


def export_and_format_excel(job: str, df: pd.DataFrame):
    """
    Exports the DataFrame to a professionally formatted Excel file with
    bold/centered headers, auto-fitted columns, and banded (alternating) rows.
    """
    if df.empty:
        print("⚠️ DataFrame is empty, skipping Excel export.")
        return

    output_filename = f"{job}_output.xlsx"
    print(f"🎨 Formatting and exporting data to Excel: {output_filename}")

    with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='FAQ_Analysis')
        
        # Get the worksheet object for formatting
        worksheet = writer.sheets['FAQ_Analysis']

        # --- 1. Create an Excel Table for banded rows and filtering ---
        # This is the modern way to get alternating row colors.
        table_range = f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"
        table = Table(displayName="FAQTable", ref=table_range)
        
        # Add a default style with banded rows and filtering enabled
        style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        worksheet.add_table(table)

        # --- 2. Format Headers and Auto-fit Columns ---
        for col_idx, column_cell in enumerate(worksheet[1], 1):
            # Make headers bold and centered
            column_cell.font = Font(bold=True)
            column_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Determine the best width for the column
            max_length = 0
            column_letter = get_column_letter(col_idx)
            
            # Check header length
            if column_cell.value:
                max_length = len(str(column_cell.value))

            # Check each cell in the column for max content length
            for cell in worksheet[column_letter]:
                try: 
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Add padding and set the final column width
            adjusted_width = (max_length + 3)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    print(f"✅ Success! File saved as '{output_filename}' in the current folder.")
